import pandas
import numpy 
import matplotlib.pyplot as paint
import xlrd
from sklearn.preprocessing import LabelBinarizer
from PyQt5 import Qt
from PyQt5.QtWidgets import QSizePolicy
from matplotlib.backends.backend_qt4agg import FigureCanvas
from PyQt5 import QtWidgets,QtGui,QtCore
from PyQt5.QtWidgets import QWidget
import matplotlib
import sklearn.cluster
import warnings
from FormForApp import str0,str01,str03,str04,str05,str06,str07,str08,str09
import sys

class Mycanvas(FigureCanvas):#класс "холст" для дендрограммы
     def __init__(self,fig):
         self.__fig=fig
         FigureCanvas.__init__(self,self.__fig)
         FigureCanvas.setSizePolicy(self,QSizePolicy.Expanding,QSizePolicy.Expanding)
         FigureCanvas.updateGeometry(self)

class Error(QWidget):# класс окна оповещения
    #str2,str1 передаются метке label и заголовку окна
     def __init__(self,str1,str2):
        super(Error, self).__init__()
        self.OK = QtWidgets.QPushButton('ОК',self)
        # кнопка ОК
        self.label = QtWidgets.QLabel(str2,self)
        # Метка, отображающая значение передаваемого параметра str2
        self.setupUi(str1)
        self.connect()
        
     def connect(self):#связывает нажатие кнопки ОK c CB_1
        self.OK.clicked.connect(self.CB_1)
    
     def setupUi(self,str1):#функция настройки параметров окна
        self.resize(480, 150)      
        self.OK.setObjectName(u"OK")
        self.OK.setGeometry(QtCore.QRect(170, 80, 110, 40))        
        self.label.setObjectName(u"label")
        self.label.setGeometry(QtCore.QRect(30, 20, 415, 60))
        self.setWindowTitle(QtCore.QCoreApplication.translate("Errordiolog", str1, None))
        QtCore.QMetaObject.connectSlotsByName(self)#в заголовок передается str1
     
     def CB_1(self):#прячет окно
         self.hide()

class Kol_N_k_sred(Error): # класс окна для введения N количества кластеров
     #str2,str1 передаются метке label и заголовку окна
     def __init__(self,str1,str2):
        super(Kol_N_k_sred, self).__init__(str1,str2)
        self.Cancel = QtWidgets.QPushButton('ОТМЕНА',self)
        self.Cancel.setObjectName(u"Cancel")
        self.Cancel.setGeometry(QtCore.QRect(250, 155, 110, 40)) 
        self.Cancel.clicked.connect(self.CB_1)
        self.setupUi(str1)
        self.connect()
     
     def connect(self):
        self.OK.clicked.connect(self.CB_2)
        
     def setupUi(self,str1):
        self.resize(480, 235)     
        self.OK.setObjectName(u"OK")
        self.OK.setGeometry(QtCore.QRect(100, 155, 110, 40))              
        self.label.setObjectName(u"label")
        self.label.setGeometry(QtCore.QRect(100, 20, 415, 60))
        self.__SpinBox = QtWidgets.QSpinBox(self)
        self.__SpinBox.setObjectName(u"SpinBox")
        self.__SpinBox.setGeometry(QtCore.QRect(190, 90, 80, 45))
        self.__SpinBox.setMinimum(1)
        self.setWindowTitle(QtCore.QCoreApplication.translate("Errordiolog", str1, None))
        self.__SpinBox.setSingleStep(1)
        QtCore.QMetaObject.connectSlotsByName(self)
     
     def CB_2(self):
       if (self.k()!=False):
         global vivod_d1 
         self.zap_t('label1',vivod_d1)

     def k(self):
        _chislo_klasterov = self.__SpinBox.value()
        global srn,std,centers,data_two,dannie_na_vivod,err2
        _do=data_two.iloc[:,1:].values
        srn=_do
        _do=(_do-_do.mean(axis=0))/_do.std(axis=0)
        std=_do
        from scipy.cluster.vq import kmeans2
        warnings.filterwarnings("error",category=UserWarning)
        try:
           centers,_labels = kmeans2(_do,_chislo_klasterov,minit='++')
        except UserWarning:
           err2.show()
           return False
        dannie_na_vivod.loc[:,'label1']=_labels

     def zap_t(self,str1,w_name):
        global dannie_na_vivod,kol
        _name_tab=[]
        w_name.tab_1.clear()
        _m=0
        for group in dannie_na_vivod.groupby(str1):
         _tab_name ="Группа №{}".format(_m+1) 
         _name_tab.append(QtWidgets.QWidget() )
         _layout = QtWidgets.QVBoxLayout(_name_tab[_m])
         _tableName = "Table_{}".format(_m) 
         _table1 = QtWidgets.QTableWidget(group[1].shape[0],dannie_na_vivod.shape[1]-1)
         _table1.setObjectName(str(_tableName))
         _table1.setHorizontalHeaderLabels(dannie_na_vivod.columns)
         for i in range(_table1.rowCount()):
            for j in range(_table1.columnCount()):
                _x = group[1].iloc[i, j]
                _table1.setItem(i,j, QtWidgets.QTableWidgetItem(str(_x)))
                if(j<=kol and j>0):
                 _table1.item(i, j).setBackground(QtGui.QColor(100, 100, 200))
         _table1.resizeColumnsToContents()
         _label = QtWidgets.QLabel('*-цветом обозначены признаки, по которым велось разбиение')
         _layout.addWidget(_label)
         _layout.addWidget(_table1)
         w_name.tab_1.addTab(_name_tab[_m], _tab_name)
         _m=_m+1
        dannie_na_vivod =dannie_na_vivod.drop(str1,1) 
        w_name.show()
           
class Dendr(Kol_N_k_sred):
     global Z,tab_1
     def __init__(self,str1,str2):
        super(Dendr, self).__init__(str1,str2)
        self.Cancel.setGeometry(QtCore.QRect(1000, 900,110, 40))
        self.setupUi(str1)
        self.connect()
        
     def setupUi(self,str1):
        self.resize(1900, 1000)
        self.verticalLayoutWidget =QWidget(self)
        self.verticalLayoutWidget.setObjectName(u"verticalLayoutWidget")
        self.verticalLayoutWidget.setGeometry(QtCore.QRect(20, 20, 1870, 800))
        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalLayoutWidget)
        self.verticalLayout.setObjectName(u"verticalLayout")
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)        
        self.label.setObjectName(u"label")
        self.label.setGeometry(QtCore.QRect(550, 850, 550, 40)) 
        self.OK.setObjectName(u"OK")
        self.OK.setGeometry(QtCore.QRect(800, 900, 110, 40))              
        self.__doubleSpinBox =QtWidgets.QDoubleSpinBox(self)
        self.__doubleSpinBox.setObjectName(u"doubleSpinBox")
        self.__doubleSpinBox.setGeometry(QtCore.QRect(1080, 850, 80, 45))
        self.__doubleSpinBox.setSingleStep(0.100000000000000)
        self.setWindowTitle(QtCore.QCoreApplication.translate("Errordiolog", str1, None))        
        QtCore.QMetaObject.connectSlotsByName(self)
     
     def CB_2(self):
        global labs
        from scipy.cluster.hierarchy import fcluster
        _chislo_klasterov = self.__doubleSpinBox.value()
        _labs=fcluster(Z,_chislo_klasterov,criterion='distance')
        labs=_labs
        print(labs)
        _b=numpy.unique(_labs)
        dannie_na_vivod.loc[:,'label']=_labs
        global vivod_d 
        self.zap_t('label',vivod_d)

class Mywindow(QtWidgets.QMainWindow):
    def __init__(self):
        global data_two,dannie_na_vivod,kol
        super(Mywindow, self).__init__()
        self.setupUi()
        self.pushButton_1.clicked.connect(self.ClickButton)
        self.pushButton_2.clicked.connect(self.ClB_l)
        self.checkBox_1.stateChanged.connect(self.clickBox_1)
        self.checkBox_2.stateChanged.connect(self.clickBox_2)
        self.checkBox_3.stateChanged.connect(self.clickBox_3)
        self.checkBox_4.stateChanged.connect(self.clickBox_4)
        self.checkBox_5.stateChanged.connect(self.clickBox_5)
        self.checkBox_6.stateChanged.connect(self.clickBox_6)
        self.checkBox_7.stateChanged.connect(self.clickBox_7)
        self.checkBox_8.stateChanged.connect(self.clickBox_8)
        self.checkBox_9.stateChanged.connect(self.clickBox_9)
        self.checkBox_10.stateChanged.connect(self.clickBox_10)
        self.checkBox_11.stateChanged.connect(self.clickBox_11)
        self.checkBox_12.stateChanged.connect(self.clickBox_12)
        self.checkBox_13.stateChanged.connect(self.clickBox_13)
        self.checkBox_14.stateChanged.connect(self.clickBox_14)
        self.checkBox_15.stateChanged.connect(self.clickBox_15)
        self.checkBox_16.stateChanged.connect(self.clickBox_16)
        self.checkBox_17.stateChanged.connect(self.clickBox_17)
        self.checkBox_18.stateChanged.connect(self.clickBox_18)
        self.checkBox_19.stateChanged.connect(self.clickBox_19)
        self.checkBox_20.stateChanged.connect(self.clickBox_20)

    def setupUi(self):
        self.resize(1250, 700)
        self.centralwidget =QWidget(self)
        self.centralwidget.setObjectName("centralwidget")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setObjectName(u"tabWidget")
        self.tabWidget.setGeometry(QtCore.QRect(20, 5, 1220, 500))
        self.tab = QWidget()
        self.tab1 = QWidget()
        self.tab.setObjectName(u"tab")
        self.tab1.setObjectName(u"tab1")
        self.label01 = QtWidgets.QLabel(str05,self.tab1)
        self.label01.setObjectName(u"label")
        self.label01.setGeometry(QtCore.QRect(20, 0, 1000, 100))
        self.label01.setFont(QtGui.QFont("Times", 13, QtGui.QFont.Bold))
        self.label02 = QtWidgets.QLabel(str06,self.tab1)
        self.label02.setObjectName(u"label")
        self.label02.setGeometry(QtCore.QRect(20, 15, 1200, 150))
        self.label03 = QtWidgets.QLabel(str07,self.tab1)
        self.label03.setObjectName(u"label")
        self.label03.setGeometry(QtCore.QRect(20, 90, 1200, 150))
        self.label04 = QtWidgets.QLabel(str08,self.tab1)
        self.label04.setObjectName(u"label")
        self.label04.setGeometry(QtCore.QRect(20, 215, 1180, 150))
        self.label05 = QtWidgets.QLabel(str09,self.tab1)
        self.label05.setObjectName(u"label")
        self.label05.setGeometry(QtCore.QRect(20, 370, 1180, 80))
        self.label05.setStyleSheet("QLabel { background-color : yellow;}");
        self.tab2 = QtWidgets.QWidget()
        self.tab2.setObjectName(u"tab2")
        self.label06 = QtWidgets.QLabel(str0,self.tab2)
        self.label06.setObjectName(u"label")
        self.label06.setGeometry(QtCore.QRect(20, 10, 1000, 100))
        self.label06.setFont(QtGui.QFont("Times", 13, QtGui.QFont.Bold))
        self.label07 = QtWidgets.QLabel(str01,self.tab2)
        self.label07.setObjectName(u"label")
        self.label07.setGeometry(QtCore.QRect(20, 50, 1200, 150))
        self.label08 = QtWidgets.QLabel(str03,self.tab2)
        self.label08.setObjectName(u"label")
        self.label08.setGeometry(QtCore.QRect(20, 150, 1200, 150))
        self.label09 = QtWidgets.QLabel(str04,self.tab2)
        self.label09.setObjectName(u"label")
        self.label09.setGeometry(QtCore.QRect(20, 280, 1180, 95))
        self.label09.setStyleSheet("QLabel { background-color : yellow;}");
        self.checkBox_1 = QtWidgets.QCheckBox( "Возраст студента",self.tab)
        self.checkBox_1.setGeometry(QtCore.QRect(150, 50, 300, 18))
        self.checkBox_1.setObjectName("checkBox")
        self.checkBox_2 = QtWidgets.QCheckBox("Наличие/отсуствие высшего образования",self.tab)
        self.checkBox_2.setGeometry(QtCore.QRect(150, 90, 400, 18))
        self.checkBox_2.setObjectName("checkBox_2")
        self.checkBox_3 = QtWidgets.QCheckBox( "Наличие/отсуствие работы",self.tab)
        self.checkBox_3.setGeometry(QtCore.QRect(150, 130, 400, 18))
        self.checkBox_3.setObjectName("checkBox_3")
        self.checkBox_4 = QtWidgets.QCheckBox("Семейное положение",self.tab)
        self.checkBox_4.setGeometry(QtCore.QRect(150, 170, 400, 18))
        self.checkBox_4.setObjectName("checkBox_4")
        self.checkBox_5 = QtWidgets.QCheckBox("Наличие/отсуствие детей",self.tab)
        self.checkBox_5.setGeometry(QtCore.QRect(150, 210, 400, 18))
        self.checkBox_5.setObjectName("checkBox_5")
        self.checkBox_6 = QtWidgets.QCheckBox( "Наличие/отсуствие родителя(ей)",self.tab)
        self.checkBox_6.setGeometry(QtCore.QRect(150, 250, 450, 18))
        self.checkBox_6.setObjectName("checkBox_6")
        self.checkBox_7 = QtWidgets.QCheckBox("Наличие/отсуствие высшего образования у родителя(ей)",self.tab)
        self.checkBox_7.setGeometry(QtCore.QRect(150, 290, 450, 18))
        self.checkBox_7.setObjectName("checkBox_7")
        self.checkBox_8 = QtWidgets.QCheckBox( "Условия проживания",self.tab)
        self.checkBox_8.setGeometry(QtCore.QRect(150, 330, 450, 18))
        self.checkBox_8.setObjectName("checkBox_8")
        self.checkBox_9 = QtWidgets.QCheckBox("Факультет",self.tab)
        self.checkBox_9.setGeometry(QtCore.QRect(150, 370, 450, 18))
        self.checkBox_9.setObjectName("checkBox_9")
        self.checkBox_10 = QtWidgets.QCheckBox("Участие в общественной жизни",self.tab)
        self.checkBox_10.setGeometry(QtCore.QRect(150, 410, 450, 18))
        self.checkBox_10.setObjectName("checkBox_10")
        self.checkBox_11 = QtWidgets.QCheckBox("Среднее количество 3-к за семестр",self.tab)
        self.checkBox_11.setGeometry(QtCore.QRect(700, 50, 450, 18))
        self.checkBox_11.setObjectName("checkBox_11")
        self.checkBox_12 = QtWidgets.QCheckBox("Среднее количество 4-к за семестр",self.tab)
        self.checkBox_12.setGeometry(QtCore.QRect(700, 90, 450, 18))
        self.checkBox_12.setObjectName("checkBox_12")
        self.checkBox_13 = QtWidgets.QCheckBox("Среднее количество 5-к за семестр",self.tab)
        self.checkBox_13.setGeometry(QtCore.QRect(700, 130, 450, 18))
        self.checkBox_13.setObjectName("checkBox_13")
        self.checkBox_14 = QtWidgets.QCheckBox("Среднее количество пересдач за семестр",self.tab)
        self.checkBox_14.setGeometry(QtCore.QRect(700, 170, 460, 18))
        self.checkBox_14.setObjectName("checkBox_14")
        self.checkBox_15 = QtWidgets.QCheckBox( "Количество пересдач с комиссией за все время обучения",self.tab)
        self.checkBox_15.setGeometry(QtCore.QRect(700, 210, 460, 18))
        self.checkBox_15.setObjectName("checkBox_15")
        self.checkBox_16 = QtWidgets.QCheckBox("Количество пропущенных в семестр лекционных занятий",self.tab)
        self.checkBox_16.setGeometry(QtCore.QRect(700, 250, 460, 18))
        self.checkBox_16.setObjectName("checkBox_16")
        self.checkBox_17 = QtWidgets.QCheckBox("Количество пропущенных в семестр практических занятий",self.tab)
        self.checkBox_17.setGeometry(QtCore.QRect(700, 290, 470, 18))
        self.checkBox_17.setObjectName("checkBox_17")
        self.checkBox_18 = QtWidgets.QCheckBox("Сроки сдачи учебных работ",self.tab)
        self.checkBox_18.setGeometry(QtCore.QRect(700, 330, 470, 18))
        self.checkBox_18.setObjectName("checkBox_18")
        self.checkBox_19 = QtWidgets.QCheckBox("Количество пересдач за все время обучения",self.tab)
        self.checkBox_19.setGeometry(QtCore.QRect(700, 370, 460, 18))
        self.checkBox_19.setObjectName("checkBox_19")
        self.checkBox_20 = QtWidgets.QCheckBox("Количество сессий, закрытых на степендию",self.tab)
        self.checkBox_20.setGeometry(QtCore.QRect(700, 410, 460, 18))
        self.checkBox_20.setObjectName("checkBox_20")
        self.tabWidget.addTab(self.tab, "Признаки для группирования")
        self.tabWidget.addTab(self.tab1, "Метод К-средних")
        self.tabWidget.addTab(self.tab2, "Иерархический метод")
        self.pushButton_1 = QtWidgets.QPushButton( "Провести анализ иерархическим методом",self.centralwidget)
        self.pushButton_1.setGeometry(QtCore.QRect(150, 550, 330, 100))
        self.pushButton_1.setObjectName("pushButton_1")
        self.pushButton_2 = QtWidgets.QPushButton( "Провести анализ методом k-средних",self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(700, 550, 330, 100))
        self.pushButton_2.setObjectName("pushButton_2")
        self.setCentralWidget(self.centralwidget)
        self.setWindowTitle(QtCore.QCoreApplication.translate("Errordiolog", "Группирование студентов", None))      
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(self)    
        
    def dob(self,str,str1,str2,str3,str4):
        global data_two,dannie_na_vivod,kol
        if (str1=='' and str2=='' and str3=='' and str4==''):
         d1=pandas.DataFrame(dannie.fit_transform(data[str]),columns=dannie.classes_)
        elif (str1!='' and str2=='' and str3=='' and str4==''):
         d1=pandas.DataFrame(dannie.fit_transform(data[str]),columns=[str1])
        elif (str1!='' and str2!='' and str3!='' and str4!=''):
         d1=pandas.DataFrame(dannie.fit_transform(data[str]),columns=[str1,str2,str3,str4])
        else: 
            if (str1=='' and str2=='kol_1' and str3=='' and str4==''):
             d1=pandas.DataFrame(data[str])
        dannie_na_vivod=dannie_na_vivod.drop(str,1)
        d=pandas.DataFrame(data[str])
        dannie_na_vivod.insert(1,str,d)
        data_two=data_two.join(d1)
        kol=kol+1
        print(data_two)

    def delete(self,str):
       global data_two,dannie_na_vivod,kol
       n=data.columns.get_loc(str)
       print(n)
       dannie_na_vivod=dannie_na_vivod.drop(str,1)
       if(n+kol-1>20):
        dannie_na_vivod.insert(19,str,data[str]) 
       else:
        dannie_na_vivod.insert(n+kol-1,str,data[str]) 
       kol=kol-1
       print(kol)

    def proverka(self):
         if (self.checkBox_1.isChecked()==False and self.checkBox_2.isChecked()==False
            and self.checkBox_3.isChecked()==False and self.checkBox_4.isChecked()==False
            and self.checkBox_5.isChecked()==False and self.checkBox_6.isChecked()==False
            and self.checkBox_7.isChecked()==False and self.checkBox_8.isChecked()==False
            and self.checkBox_9.isChecked()==False and self.checkBox_10.isChecked()==False
            and self.checkBox_11.isChecked()==False and self.checkBox_12.isChecked()==False
            and self.checkBox_13.isChecked()==False and self.checkBox_14.isChecked()==False
            and self.checkBox_15.isChecked()==False and self.checkBox_16.isChecked()==False
            and self.checkBox_17.isChecked()==False and self.checkBox_18.isChecked()==False
            and self.checkBox_19.isChecked()==False and self.checkBox_20.isChecked()==False):
             return False 
         else:
             return True

    def ClB_l(self):
        if self.proverka()==False: 
         err.show()  
         return
        test.show()
        global dat_nov
        dat_nov=data_two.loc[1:1]
        print(dat_nov)

    def ClickButton(self):
        global kol
        if self.proverka()==False: 
         err.show()  
         return
        if kol==1: 
         err3.show()  
         return
        global dat_nov,srn,std,figu,data_two,canvas,Z,test2
        self.Ier_a()
    
    def Ier_a(self):
        global dat_nov,srn,std,figu,data_two,canvas,Z,test2
        dat_nov=data_two.loc[1:1]
        _do=data_two.iloc[:,1:].values
        srn=_do
        _do=(_do-_do.mean(axis=0))/_do.std(axis=0)
        std=_do
        from scipy.cluster.hierarchy import linkage,fcluster,dendrogram
        Z=linkage(_do,method='ward',metric='euclidean') 
        figu=matplotlib.pyplot.figure()
        dendogr=dendrogram(Z,orientation='top',labels=data_two.ID_студента.values)
        test2.verticalLayout.removeWidget(canvas)
        canvas=Mycanvas(figu)
        test2.verticalLayout.addWidget(canvas)
        test2.show()
   
    def clickBox_1(self):
       if self.checkBox_1.isChecked():
            self.dob('Возраст','','','','')
       else:
            global data_two
            data_two =data_two.drop('20-25',1) 
            data_two =data_two.drop('15-19',1)
            data_two =data_two.drop('больше 30',1)
            self.delete('Возраст')
           
    def clickBox_2(self):
       if self.checkBox_2.isChecked():
            self.dob('Первое_высшее_образование','Первое высшее образование(0-да,1-нет)','','','')  
       else:
            global data_two
            data_two =data_two.drop('Первое высшее образование(0-да,1-нет)',1)
            self.delete('Первое_высшее_образование')
           
    def clickBox_3(self):
       if self.checkBox_3.isChecked():
            self.dob('Работа','Работа(1-есть,0-нет)','','','')
       else:
            global data_two
            data_two =data_two.drop('Работа(1-есть,0-нет)',1)      
            self.delete('Работа')
    
    def clickBox_4(self):
       if self.checkBox_4.isChecked():
            self.dob('Семейное_положение','1-не женат/не замужем 0-женат/замужем','','','')
       else:
            global data_two
            data_two =data_two.drop('1-не женат/не замужем 0-женат/замужем',1)
            self.delete('Семейное_положение')

    def clickBox_5(self):
       if self.checkBox_5.isChecked():
            self.dob('Дети','Есть дети (0-да 1-нет)','','','')
       else:
            global data_two
            data_two =data_two.drop('Есть дети (0-да 1-нет)',1)
            self.delete('Дети')

    def clickBox_6(self):
       if self.checkBox_6.isChecked():
            self.dob('Оба_родителя','','','','') 
       else:
            global data_two
            data_two =data_two.drop('в основном с одним, т.к. второго видел нечасто',1)
            data_two =data_two.drop('да',1)
            data_two =data_two.drop('нет, я рос без родителей',1)
            data_two =data_two.drop('нет, я рос с одним родителем',1)
            self.delete('Оба_родителя')

    def clickBox_7(self):
       if self.checkBox_7.isChecked():
            self.dob('Наличие_образования_у_родителей','','','','') 
       else:
            global data_two
            data_two =data_two.drop('нет высшего образования, но есть среднее профессиональное',1) 
            data_two =data_two.drop('нет высшего образования',1)
            data_two =data_two.drop('есть высшее образование',1)
            self.delete('Наличие_образования_у_родителей')

    def clickBox_8(self):
      if self.checkBox_8.isChecked():
            self.dob('Условия_проживания','','','','')
      else:
            global data_two
            data_two =data_two.drop('вместе со своими родителями',1) 
            data_two =data_two.drop('в отдельной квартире',1)
            data_two =data_two.drop('в общежитии',1)
            self.delete('Условия_проживания')

    def clickBox_9(self):
       if self.checkBox_9.isChecked():
            self.dob('Факультет','','','','')
       else:
            global data_two
            data_two =data_two.drop('АВТИ',1) 
            data_two =data_two.drop('ГПИ',1)
            data_two =data_two.drop('ИГВИЭ',1)
            data_two =data_two.drop('ИПЭЭФ',1)
            data_two =data_two.drop('ИРЭ',1)
            data_two =data_two.drop('ИТАЭ',1)
            data_two =data_two.drop('ИЭТ',1)
            data_two =data_two.drop('ИЭЭ',1)
            data_two =data_two.drop('ИнЭИ',1)
            data_two =data_two.drop('ЭнМИ',1)
            self.delete('Факультет')
           
    def clickBox_10(self):
       if self.checkBox_10.isChecked():
            self.dob('Участие_в_общественной_жизни_университета','','','','')
       else:
            global data_two
            data_two =data_two.drop('редко',1) 
            data_two =data_two.drop('довольно часто',1)
            data_two =data_two.drop('время от времени',1)
            self.delete('Участие_в_общественной_жизни_университета')

    def clickBox_11(self):
       if self.checkBox_11.isChecked():
            self.dob('Среднее_количество_3-к','','kol_1','','')
       else:
            global data_two
            data_two =data_two.drop('Среднее_количество_3-к',1)
            self.delete('Среднее_количество_3-к')

    def clickBox_12(self):
       if self.checkBox_12.isChecked():
            self.dob('Среднее_количество_4-к','','kol_1','','')
       else:
            global data_two 
            data_two =data_two.drop('Среднее_количество_4-к',1)
            self.delete('Среднее_количество_4-к')

    def clickBox_13(self):
       if self.checkBox_13.isChecked():
           self.dob('Среднее_количество_5-к','','kol_1','','')
       else:
            global data_two
            data_two =data_two.drop('Среднее_количество_5-к',1)
            self.delete('Среднее_количество_5-к')

    def clickBox_14(self):
       if self.checkBox_14.isChecked():
            self.dob('Среднее_количество_передач_за_семестр','','kol_1','','')
       else:
            global data_two
            data_two =data_two.drop('Среднее_количество_передач_за_семестр',1)
            self.delete('Среднее_количество_передач_за_семестр')

    def clickBox_15(self):
       if self.checkBox_15.isChecked():
            self.dob('Количество_пересдач_с_коммиссией','','kol_1','','')
       else:
            global data_two
            data_two =data_two.drop('Количество_пересдач_с_коммиссией',1)
            self.delete('Количество_пересдач_с_коммиссией')

    def clickBox_16(self):
       if self.checkBox_16.isChecked():
            self.dob('Семестровое_количество_пропущенных_лекций','','','','')
       else:
            global data_two
            data_two =data_two.drop('0-3',1)
            data_two =data_two.drop('4-7',1)
            data_two =data_two.drop('7-15',1)
            data_two =data_two.drop('более 15',1)
            self.delete('Семестровое_количество_пропущенных_лекций')

    def clickBox_17(self):
       if self.checkBox_17.isChecked():
            self.dob('Семестровое_количество_пропущенных_занятий','a','b','c','d')
       else:
            global data_two
            data_two =data_two.drop('a',1)
            data_two =data_two.drop('b',1)
            data_two =data_two.drop('c',1)
            data_two =data_two.drop('d',1)
            self.delete('Семестровое_количество_пропущенных_занятий')

    def clickBox_18(self):
       if self.checkBox_18.isChecked():
            self.dob('Сроки_сдачи_учебных_работ','','','','')
       else:
            global data_two
            data_two =data_two.drop('скорее часть работ сдаю в течении семестра, часть оставляю на конец',1)
            data_two =data_two.drop('скорее сдаю работы в течении семестра , как и предусмотрено учебным планом',1)
            data_two =data_two.drop('все работы сдаю в конце семестра',1)
            data_two =data_two.drop('скорее сдаю все заранее',1)
            self.delete('Сроки_сдачи_учебных_работ')
    
    def clickBox_19(self):
       if self.checkBox_19.isChecked():
            self.dob('Среднее_количество_пересдач_за_все_время_обучения','','kol_1','','')
       else:
            global data_two
            data_two =data_two.drop('Среднее_количество_пересдач_за_все_время_обучения',1)
            self.delete('Среднее_количество_пересдач_за_все_время_обучения')
    
    def clickBox_20(self):
       if self.checkBox_20.isChecked():
            self.dob('Количество_сессий_закрытых_на_степендию','','kol_1','','')
       else:
            global data_two
            data_two =data_two.drop('Количество_сессий_закрытых_на_степендию',1)
            self.delete('Количество_сессий_закрытых_на_степендию')

class Myvivod(QtWidgets.QWidget):
    def __init__(self,str):
        super(Myvivod, self).__init__()
        self.setupUi(str)
        if (str=="Метод к-средних"):
         self.b1.clicked.connect(self.ClickButton)
        if (str=="Иерархический метод"):
         self.b1.clicked.connect(self.ClickButton1)
        self.b2.clicked.connect(self.sohr)
    def setupUi(self,str):
        self.resize(1250, 900)
        self.tab_1=QtWidgets.QTabWidget(self)
        self.tab_1.setWindowTitle(str)
        self.tab_1.setGeometry(QtCore.QRect(20, 5, 1220, 500))
        self.vvod_d=QWidget(self)
        self.vvod_d.setGeometry(QtCore.QRect(20, 550, 1220, 150))
        self.layout_d = QtWidgets.QVBoxLayout(self.vvod_d)
        self.table1 = QtWidgets.QTableWidget(1,21)
        self.table1.setHorizontalHeaderLabels(data.columns)
        self.table1.resizeColumnsToContents()
        from openpyxl import load_workbook
        book = load_workbook("Answers.xlsx")
        b1 = book.get_sheet_by_name("answers")
        i = 1
        while b1.cell(row=i, column=1).value is not None:
         i=i+1
        self.table1.setItem(0,0, QtWidgets.QTableWidgetItem("Студент №{}".format(i-1)))
        self.layout_d.addWidget(self.table1)
        self.b1=QtWidgets.QPushButton('Прогноз',self)
        self.b1.setGeometry(QtCore.QRect(230, 750, 150, 60)) 
        self.b2=QtWidgets.QPushButton('Сохранение',self)
        self.b2.setGeometry(QtCore.QRect(30, 750, 150, 60))
        self.setWindowTitle(QtCore.QCoreApplication.translate("Errordiolog", str, None))  
        self.label = QtWidgets.QLabel('Поле для записи информации о новом студенте',self)
        self.label.setGeometry(QtCore.QRect(30, 510, 415, 60))
        self.label1 = QtWidgets.QLabel('Группа',self)   
        self.label1.setGeometry(QtCore.QRect(450, 750, 95, 60))
        self.label1.setStyleSheet("QLabel { background-color : yellow;}");
        self.cb1 = QtWidgets.QComboBox()
        self.cb1.addItems(["больше 30","15-19", "20-25"])
        self.table1.setCellWidget(0,1,self.cb1)
        self.cb2 = QtWidgets.QComboBox()
        self.cb2.addItems(["да", "нет"])
        self.table1.setCellWidget(0,2,self.cb2)
        self.cb3 = QtWidgets.QComboBox()
        self.cb3.addItems(["да", "нет"])
        self.table1.setCellWidget(0,3,self.cb3)
        self.cb4 = QtWidgets.QComboBox()
        self.cb4.addItems(["женат/замужем", "не женат/не замужем"])
        self.table1.setCellWidget(0,4,self.cb4)
        self.cb5 = QtWidgets.QComboBox()
        self.cb5.addItems(["да", "нет"])
        self.table1.setCellWidget(0,5,self.cb5)
        self.cb6 = QtWidgets.QComboBox()
        self.cb6.addItems(["нет, я рос с одним родителем","да","в основном с одним, т.к. второго видел нечасто","нет, я рос без родителей"])
        self.table1.setCellWidget(0,6,self.cb6)
        self.cb7 = QtWidgets.QComboBox()
        self.cb7.addItems(["нет высшего образования, но есть среднее профессиональное","есть высшее образование", "нет высшего образования"])
        self.table1.setCellWidget(0,7,self.cb7)
        self.cb8 = QtWidgets.QComboBox()
        self.cb8.addItems(["вместе со своими родителями","в общежитии", "в отдельной квартире"])
        self.table1.setCellWidget(0,8,self.cb8)
        self.cb9 = QtWidgets.QComboBox()
        self.cb9.addItems(["АВТИ","ЭнМИ", "ИПЭЭФ", "ИЭТ", "ИЭЭ", "ИРЭ", "ГПИ", "ИнЭИ"\
            , "ИГВИЭ", "ИТАЭ"])
        self.table1.setCellWidget(0,9,self.cb9)
        self.cb10 = QtWidgets.QComboBox()
        self.cb10.addItems(["довольно часто","редко", "время от времени"])
        self.table1.setCellWidget(0,10,self.cb10)
        self.cb11 = QtWidgets.QSpinBox()
        self.cb11.setMinimum(0)
        self.cb11.setMaximum(10)
        self.cb11.setSingleStep(1)
        self.table1.setCellWidget(0,11,self.cb11)
        self.cb12 = QtWidgets.QSpinBox()
        self.cb12.setMinimum(0)
        self.cb12.setMaximum(10)
        self.cb12.setSingleStep(1)
        self.table1.setCellWidget(0,12,self.cb12)
        self.cb13 = QtWidgets.QSpinBox()
        self.cb13.setMinimum(0)
        self.cb13.setMaximum(10)
        self.cb13.setSingleStep(1)
        self.table1.setCellWidget(0,13,self.cb13)
        self.cb14 = QtWidgets.QSpinBox()
        self.cb14.setMinimum(0)
        self.cb14.setMaximum(5)
        self.cb14.setSingleStep(1)
        self.table1.setCellWidget(0,14,self.cb14)
        self.cb15 = QtWidgets.QSpinBox()
        self.cb15.setMinimum(0)
        self.cb15.setMaximum(5)
        self.cb15.setSingleStep(1)
        self.table1.setCellWidget(0,15,self.cb15)
        self.cb16 = QtWidgets.QComboBox()
        self.cb16.addItems(["0-3","4-7", "7-15","более 15"])
        self.table1.setCellWidget(0,16,self.cb16)
        self.cb17 = QtWidgets.QComboBox()
        self.cb17.addItems(["0-3","4-7", "7-15","более 15"])
        self.table1.setCellWidget(0,17,self.cb17)
        self.cb18 = QtWidgets.QComboBox()
        self.cb18.addItems(["скорее сдаю работы в течении семестра , как и предусмотрено учебным планом",\
            "скорее часть работ сдаю в течении семестра, часть оставляю на конец",\
           "скорее сдаю все заранее","все работы сдаю в конце семестра"])
        self.table1.setCellWidget(0,18,self.cb18)
        self.cb19 = QtWidgets.QSpinBox()
        self.cb19.setMinimum(0)
        self.cb19.setMaximum(10)
        self.cb19.setSingleStep(1)
        self.table1.setCellWidget(0,19,self.cb19)
        self.cb20 = QtWidgets.QSpinBox()
        self.cb20.setMinimum(0)
        self.cb20.setMaximum(6)
        self.cb20.setSingleStep(1)
        self.table1.setCellWidget(0,20,self.cb20)

    def ClickButton(self):
           global dat_nov
           self.chit()
           print(dat_nov)
           global srn,std,centers
           _do=dat_nov.iloc[:,1:].values
           _do=(_do-srn.mean(axis=0))/srn.std(axis=0)
           self.funk_k(_do,centers)

    def ClickButton1(self):
           global dat_nov
           self.chit()
           print(dat_nov)
           global srn,std,labs
           _do=dat_nov.iloc[:,1:].values
           _do=(_do-srn.mean(axis=0))/srn.std(axis=0)
           self.funk_i(_do,std,labs)
    
    def chit(self):
        global dat_nov
        dat_nov[0:21]=0
        text1 = self.table1.cellWidget(0,1).currentText()
        text2 = self.table1.cellWidget(0,2).currentText()
        text3 = self.table1.cellWidget(0,3).currentText()
        text4 = self.table1.cellWidget(0,4).currentText()
        text5 = self.table1.cellWidget(0,5).currentText()
        text6 = self.table1.cellWidget(0,6).currentText()
        text7 = self.table1.cellWidget(0,7).currentText()
        text8 = self.table1.cellWidget(0,8).currentText()
        text9 = self.table1.cellWidget(0,9).currentText()
        text10 = self.table1.cellWidget(0,10).currentText()
        text11 = self.table1.cellWidget(0,11).value()
        text12 = self.table1.cellWidget(0,12).value()
        text13 = self.table1.cellWidget(0,13).value()
        text14 = self.table1.cellWidget(0,14).value()
        text15 = self.table1.cellWidget(0,15).value()
        text16 = self.table1.cellWidget(0,16).currentText()
        text17 = self.table1.cellWidget(0,17).currentText()
        text18 = self.table1.cellWidget(0,18).currentText()
        text19 = self.table1.cellWidget(0,19).value()
        text20 = self.table1.cellWidget(0,20).value()
       
        if text1  in dat_nov.columns:
            dat_nov[text1]=1
        if 'Первое высшее образование(0-да,1-нет)'  in dat_nov.columns:
            if  (text2=='нет'):
             dat_nov['Первое высшее образование(0-да,1-нет)']=1
        if 'Работа(1-есть,0-нет)'  in dat_nov.columns:
            if  (text3=='да'):
             dat_nov['Работа(1-есть,0-нет)']=1
        if  '1-не женат/не замужем 0-женат/замужем' in dat_nov.columns: 
            if  (text4=='не женат/не замужем'):
                dat_nov['1-не женат/не замужем 0-женат/замужем']=1
        if  'Есть дети (0-да 1-нет)' in dat_nov.columns:
            if  (text5=='нет'):
                dat_nov['Есть дети (0-да 1-нет)']=1
        if text6  in dat_nov.columns:
            dat_nov[text6]=1
        if text7  in dat_nov.columns:
            dat_nov[text7]=1
        if text8  in dat_nov.columns:
            dat_nov[text8]=1
        if text9  in dat_nov.columns:
            dat_nov[text9]=1
        if text10  in dat_nov.columns:
            dat_nov[text10]=1
        if  'Среднее_количество_3-к' in dat_nov.columns:
            dat_nov['Среднее_количество_3-к']=text11
        if  'Среднее_количество_4-к' in dat_nov.columns:
            dat_nov['Среднее_количество_4-к']=text12
        if  'Среднее_количество_5-к' in dat_nov.columns:
            dat_nov['Среднее_количество_5-к']=text13
        if  'Среднее_количество_передач_за_семестр' in dat_nov.columns:
            dat_nov['Среднее_количество_передач_за_семестр']=text14
        if  'Количество_пересдач_с_коммиссией' in dat_nov.columns:
            dat_nov['Количество_пересдач_с_коммиссией']=text15
        if text16  in dat_nov.columns:
            dat_nov[text16]=1 
        if  'a' and 'b' and 'c' and 'd' in dat_nov.columns:
           if  (text17=='0-3'):
               dat_nov['a']=1 
           if  (text17=='4-7'):
               dat_nov['b']=1
           if  (text17=='7-15'):
               dat_nov['c']=1
           if  (text17=='более 15'):
               dat_nov['d']=1
        if text18  in dat_nov.columns:
            dat_nov[text18]=1 
        if  'Среднее_количество_пересдач_за_все_время_обучения' in dat_nov.columns:
            dat_nov['Среднее_количество_пересдач_за_все_время_обучения']=text19
        if  'Количество_сессий_закрытых_на_степендию' in dat_nov.columns:
            dat_nov['Количество_сессий_закрытых_на_степендию']=text20
               
    def funk_k(self,do,centers):
        from scipy.spatial import distance
        min=distance.euclidean(centers[0], do)
        m=0
        gr=0
        for clast in centers:
         m=m+1
         d = distance.euclidean(clast, do)
         if (d<min):
          min=d
          gr=m
         if (gr==0):
          gr=1
        self.label1.setText("Группа №{}".format(gr) ) 
      
    def funk_i(self,do,std,labs):  
        from scipy.spatial import distance
        min=distance.euclidean(std[0], do)
        m=0
        gr=0
        for clast in std:
          d = distance.euclidean(clast, do)
          if (d<min):
           min=d
           gr=labs[m]
          m=m+1
        if (gr==0):
            gr=labs[0]
        self.label1.setText("Группа №{}".format(gr) ) 
            
    def sohr(self): 
        from openpyxl import load_workbook
        book = load_workbook("Answers.xlsx")
        b1 = book.get_sheet_by_name("answers")
        i = 1
        while b1.cell(row=i, column=1).value is not None:
         i=i+1
        b1.cell(row=i, column=1).value = self.table1.cellWidget(0,0).value()
        b1.cell(row=i, column=2).value = self.table1.cellWidget(0,1).currentText()
        b1.cell(row=i, column=3).value= self.table1.cellWidget(0,2).currentText()
        b1.cell(row=i, column=4).value = self.table1.cellWidget(0,3).currentText()
        b1.cell(row=i, column=5).value= self.table1.cellWidget(0,4).currentText()
        b1.cell(row=i, column=6).value = self.table1.cellWidget(0,5).currentText()
        b1.cell(row=i, column=7).value = self.table1.cellWidget(0,6).currentText()
        b1.cell(row=i, column=8).value= self.table1.cellWidget(0,7).currentText()
        b1.cell(row=i, column=9).value= self.table1.cellWidget(0,8).currentText()
        b1.cell(row=i, column=10).value = self.table1.cellWidget(0,9).currentText()
        b1.cell(row=i, column=11).value = self.table1.cellWidget(0,10).currentText()
        b1.cell(row=i, column=12).value= self.table1.cellWidget(0,11).value()
        b1.cell(row=i, column=13).value = self.table1.cellWidget(0,12).value()
        b1.cell(row=i, column=14).value = self.table1.cellWidget(0,13).value()
        b1.cell(row=i, column=15).value = self.table1.cellWidget(0,14).value()
        b1.cell(row=i, column=16).value = self.table1.cellWidget(0,15).value()
        b1.cell(row=i, column=17).value = self.table1.cellWidget(0,16).currentText()
        b1.cell(row=i, column=18).value = self.table1.cellWidget(0,17).currentText()
        b1.cell(row=i, column=19).value = self.table1.cellWidget(0,18).currentText()
        b1.cell(row=i, column=20).value = self.table1.cellWidget(0,19).value()
        b1.cell(row=i, column=21).value = self.table1.cellWidget(0,20).value()
        book.save("Answers.xlsx")

if __name__ == "__main__": #начало основной программы
  
 str1="Ошибка №1"
 str2="Вы не выбрали ни одного признака для группирования.\n        Пожалуйста, выберите хотя бы один признак!"
 str3="Метод к-средних"
 str4="Введите количество N кластеров."
 str5="Ошибка №2"
 str6="  Один или более кластеров могут оказаться пустыми.\n  Пожалуйста, выберите меньшее число N кластеров!"   
 str7="Иерархический метод"
 str8="Выберете порог отсечения дендрограммы для получения разбиения!"
 str9="Иерархический метод"
 str10="Метод к-средних"
 #переменные для хранения заголовков или надпесей окон приложения
 Z=0 
 #переменная для хранения матрицы связи
 kol=0 
 #переменная для хранения количества выбранных признаков
 figu=matplotlib.pyplot.figure()
 #переменная для хранения дендрограммы
 data = pandas.read_excel('C:\\Users\\Nast\\source\\repos\\Test11\\Answers.xlsx')
 #таблица исходных данных
 dannie=LabelBinarizer()
 #класс для кодирования данных
 data_two=pandas.DataFrame(data['ID_студента'])
 #таблица для записи данных о выбранных признаках
 dannie_na_vivod=pandas.DataFrame(data)
 #данные на вывод
 dat_nov=0
 #переменная для хранения данных о новом студенте
 srn=0
 #переменная для хранения массива значений признаков
 std=0
 #переменная для хранения массива стандартизированных значений признаков
 centers=0
 #переменная для хранения центроидов кластеров
 labs=0
 #переменная для хранения массива меток кластеров

 app = QtWidgets.QApplication([]) 

 err=Error(str1,str2)
 err2=Error(str5,str6)
 #экземпляры класса окна-оповещения  
 
 test=Kol_N_k_sred(str3,str4)
 #экземпляр класса окна для введения N числа кластеров
 
 test2=Dendr(str7,str8)
 #экземпляр класса окна, содержащего дендрограмму

 canvas=Mycanvas(figu)
 #экземпляр класса "холста" для дендрограммы
 
 vivod_d=Myvivod(str9)
 vivod_d1 =Myvivod(str10)
 #экземпляры класса окна для отображения кластеров 
 
 application = Mywindow()
 #экземпляр класса главного окна приложеня
 
 application.show()
 
 sys.exit(app.exec()) #конец основной программы





