import pandas
import numpy 
import matplotlib.pyplot as paint
import xlrd
from sklearn.preprocessing import LabelBinarizer
from PyQt5 import Qt
from PyQt5.QtWidgets import QSizePolicy
from matplotlib.backends.backend_qt4agg import FigureCanvas
from PyQt5 import QtWidgets,QtGui,QtCore
from PyQt5.QtWidgets import QWidget
import matplotlib
import warnings
from FormForApp import str0,str01,str03,str04,str05,str06,str07,str08,str09
import sys

def data1():
        #функция для получения числовых значений выбранных признаков
        # c помощью списка L какие признаки были выбраны пользователем
        # на основе полученной информации формирует таблицу со значениями соответствующих признаков
        # возвращает массив числовых значений выбранных признаков
        global L,data,dat_preob
        copy_dat=pandas.DataFrame(data['ID_студента'])
        if 'Возраст'  in L:
           copy_dat=copy_dat.join(dat_preob.iloc[:,1:4])
        if 'Первое_высшее_образование'  in L:
           copy_dat=copy_dat.join(dat_preob['Первое_высшее_образование_преоб'])
        if 'Работа'  in L:
           copy_dat=copy_dat.join(dat_preob['Работа_преоб'])
        if 'Семейное_положение'  in L:
            copy_dat=copy_dat.join(dat_preob['Семейное_положение_преоб'])
        if 'Дети'  in L:
            copy_dat=copy_dat.join(dat_preob['Дети_преоб'])
        if 'Оба_родителя'  in L:
           copy_dat=copy_dat.join(dat_preob.iloc[:,8:12])
        if 'Наличие_образования_у_родителей'  in L:
            copy_dat=copy_dat.join(dat_preob.iloc[:,12:15])
        if 'Условия_проживания'  in L:
            copy_dat=copy_dat.join(dat_preob.iloc[:,15:18])
        if 'Факультет'  in L:
            copy_dat=copy_dat.join(dat_preob.iloc[:,18:28])
        if 'Участие_в_общественной_жизни_университета'  in L:
            copy_dat=copy_dat.join(dat_preob.iloc[:,28:31])
        if 'Среднее_количество_3-к'  in L:
            copy_dat=copy_dat.join(data['Среднее_количество_3-к'])  
        if 'Среднее_количество_4-к'  in L:
            copy_dat=copy_dat.join(data['Среднее_количество_4-к']) 
        if 'Среднее_количество_5-к'  in L:
            copy_dat=copy_dat.join(data['Среднее_количество_5-к']) 
        if 'Среднее_количество_передач_за_семестр'  in L:
           copy_dat=copy_dat.join(data['Среднее_количество_передач_за_семестр']) 
        if 'Количество_пересдач_с_коммиссией'  in L:
            copy_dat=copy_dat.join(data['Количество_пересдач_с_коммиссией']) 
        if 'Семестровое_количество_пропущенных_лекций'  in L:
            copy_dat=copy_dat.join(dat_preob.iloc[:,31:35])
        if  'Семестровое_количество_пропущенных_занятий' in L:
            copy_dat=copy_dat.join(dat_preob.iloc[:,35:39])
        if  'Сроки_сдачи_учебных_работ' in L:
            copy_dat=copy_dat.join(dat_preob.iloc[:,39:])
        if  'Среднее_количество_пересдач_за_все_время_обучения' in L:
            copy_dat=copy_dat.join(data['Среднее_количество_пересдач_за_все_время_обучения']) 
        if  'Количество_сессий_закрытых_на_степендию' in L:
            copy_dat=copy_dat.join(data['Количество_сессий_закрытых_на_степендию']) 
        dat=copy_dat.iloc[:,1:].values   
        return dat

class Mycanvas(FigureCanvas):#класс "холст" для дендрограммы
     def __init__(self,fig):
         self.__fig=fig
         FigureCanvas.__init__(self,self.__fig)
         FigureCanvas.setSizePolicy(self,QSizePolicy.Expanding,QSizePolicy.Expanding)
         FigureCanvas.updateGeometry(self)

class Error(QWidget):# класс окна оповещения
    #str2,str1 передаются метке label и заголовку окна
     def __init__(self,str1,str2):
        super(Error, self).__init__()
        self.OK = QtWidgets.QPushButton('ОК',self)
        # кнопка ОК
        self.label = QtWidgets.QLabel(str2,self)
        # Метка, отображающая значение передаваемого параметра str2
        self.setupUi(str1)
        self.connect()
        
     def connect(self):#связывает нажатие кнопки ОK c CB_1
        self.OK.clicked.connect(self.CB_1)
    
     def setupUi(self,str1):#настройка параметров окна
        self.resize(480, 150)      
        self.OK.setObjectName(u"OK")
        self.OK.setGeometry(QtCore.QRect(170, 80, 110, 40))        
        self.label.setObjectName(u"label")
        self.label.setGeometry(QtCore.QRect(30, 20, 415, 60))
        self.setWindowTitle(QtCore.QCoreApplication.translate("Errordiolog", str1, None))
        QtCore.QMetaObject.connectSlotsByName(self)#в заголовок передается str1
     
     def CB_1(self):#прячет окно
         self.hide()

class Kol_N_k_sred(Error): # класс окна для введения m количества кластеров
     #наследует атрибуты и методы класса Error
     #str2,str1 передаются метке label и заголовку окна
     def __init__(self,str1,str2):
        super(Kol_N_k_sred, self).__init__(str1,str2)
        self.Cancel = QtWidgets.QPushButton('ОТМЕНА',self)
        #кнопка ОТМЕНА
        self.Cancel.setObjectName(u"Cancel")
        self.Cancel.setGeometry(QtCore.QRect(250, 155, 110, 40)) 
        self.Cancel.clicked.connect(self.CB_1)
        #связывает нажатие кнопки ОТМЕНА с функцией CB_1 родительского класса Error
        self.setupUi(str1)
        self.connect()
     
     def connect(self):#связывает нажатие кнопки ОK c CB_2
        self.OK.clicked.connect(self.CB_2)
        
     def setupUi(self,str1):#настройка параметров окна
        self.resize(480, 235)     
        self.OK.setObjectName(u"OK")
        self.OK.setGeometry(QtCore.QRect(100, 155, 110, 40))              
        self.label.setObjectName(u"label")
        self.label.setGeometry(QtCore.QRect(100, 20, 415, 60))
        self.__SpinBox = QtWidgets.QSpinBox(self)
        #поле для введения N количества кластеров
        self.__SpinBox.setObjectName(u"SpinBox")
        self.__SpinBox.setGeometry(QtCore.QRect(190, 90, 80, 45))
        self.__SpinBox.setMinimum(1)
        self.setWindowTitle(QtCore.QCoreApplication.translate("Errordiolog", str1, None))
        self.__SpinBox.setSingleStep(1)
        QtCore.QMetaObject.connectSlotsByName(self)
     
     def CB_2(self):
       #содержит метод группирования и метод отображения результатов в новом окне
       if (self.k()!=False):
       #если группирование прошло успешно vivod_d1 отображает полученные кластеры 
         global vivod_d1 #экземпляр класса окна вывода
         self.zap_t('label1',vivod_d1)

     def k(self):#функция группирования методом k-средних++
        _chislo_klasterov = self.__SpinBox.value()
        #считывается N количество кластеров
        global srn,std,centers,err2,data
        _do=data1()
        #загружаются значения признаков без идентификаторов студентов
        srn=_do
        #массив значений запоминается
        _do=(_do-_do.mean(axis=0))/_do.std(axis=0)
        #данные стандартизируются
        std=_do
        #запоминаем стандартизированные данные
        from scipy.cluster.vq import kmeans2
        warnings.filterwarnings("error",category=UserWarning)
        try:
           centers,_labels = kmeans2(_do,_chislo_klasterov,minit='++')
           #производится кластеризация,в centers записываются координаты центров масс
           #в _labels записываются метки кластеров
        except UserWarning:
           err2.show()
           #при возникновении предупреждения появляется окно оповещение об ошибке
           #и функция возвращает False
           return False
        data.loc[:,'label1']=_labels
        #каждому студенту присваивается метка кластера
     
     def zap_t(self,str1,w_name):#процедура отображения полученных кластеров
        #str1-имя столбца меток кластеров
        #w_name-имя экземпляра класса окна для отображения результатов 
        global data,L
        _name_tab=[]
        w_name.tab_1.clear()
        _m=0
        for group in data.groupby(str1):
         #для каждого кластера создается новая вкладка и новая таблица, содержащая 
         #информацию о студентах и соответствующих им признаках 
         #каждая таблица помещается во вкладку, которая добавляется в окно для вывода результатов
         _tab_name ="Группа №{}".format(_m+1) 
         _name_tab.append(QtWidgets.QWidget() )
         _layout = QtWidgets.QVBoxLayout(_name_tab[_m])
         _tableName = "Table_{}".format(_m) 
         _table1 = QtWidgets.QTableWidget(group[1].shape[0],data.shape[1]-1)
         _table1.setObjectName(str(_tableName))
         _table1.setHorizontalHeaderLabels(data.columns)
         for i in range(_table1.rowCount()):#поэлементное заполнение таблицы данными о студентах
            for j in range(_table1.columnCount()):
                _x = group[1].iloc[i, j]
                _table1.setItem(i,j, QtWidgets.QTableWidgetItem(str(_x)))
                if(data.columns[j] in L):
                 _table1.item(i, j).setBackground(QtGui.QColor(100, 100, 200))
         _table1.resizeColumnsToContents()
         _label = QtWidgets.QLabel('*-цветом обозначены признаки, по которым велось разбиение')
         _layout.addWidget(_label)
         _layout.addWidget(_table1)
         w_name.tab_1.addTab(_name_tab[_m], _tab_name)
         _m=_m+1
        data =data.drop(str1,1) 
        w_name.show()
           
class Dendr(Kol_N_k_sred):#класс окна,содержащего дендрограмму
    #наследует атрибуты и методы класса Kol_N_k_sred
     #str2,str1 передаются метке label и заголовку окна
     global Z,tab_1
     def __init__(self,str1,str2):
        super(Dendr, self).__init__(str1,str2)
        self.Cancel.setGeometry(QtCore.QRect(1000, 900,110, 40))
        self.setupUi(str1)
        self.connect()
        
     def setupUi(self,str1):#настройка параметров окна
        self.resize(1900, 1000)
        self.verticalLayoutWidget =QWidget(self)
        self.verticalLayoutWidget.setObjectName(u"verticalLayoutWidget")
        self.verticalLayoutWidget.setGeometry(QtCore.QRect(20, 20, 1870, 800))
        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalLayoutWidget)
        self.verticalLayout.setObjectName(u"verticalLayout")
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)        
        self.label.setObjectName(u"label")
        self.label.setGeometry(QtCore.QRect(550, 850, 550, 40)) 
        self.OK.setObjectName(u"OK")
        self.OK.setGeometry(QtCore.QRect(800, 900, 110, 40))              
        self.__doubleSpinBox =QtWidgets.QDoubleSpinBox(self)
        #поле для введения порога отсечения дендрограммы
        self.__doubleSpinBox.setObjectName(u"doubleSpinBox")
        self.__doubleSpinBox.setGeometry(QtCore.QRect(1080, 850, 80, 45))
        self.__doubleSpinBox.setSingleStep(0.100000000000000)
        self.setWindowTitle(QtCore.QCoreApplication.translate("Errordiolog", str1, None))        
        QtCore.QMetaObject.connectSlotsByName(self)
     
     def CB_2(self):
        #осущевствляет группирование объектов по полученным меткам кластеров
        #отображения полученных кластеров
        global labs
        from scipy.cluster.hierarchy import fcluster
        _chislo_klasterov = self.__doubleSpinBox.value()
        _labs=fcluster(Z,_chislo_klasterov,criterion='distance')
        #получение массива меток кластеров
        labs=_labs
        _b=numpy.unique(_labs)
        data.loc[:,'label']=_labs
        global vivod_d 
        self.zap_t('label',vivod_d)#отображения полученных кластеров

class Mywindow(QtWidgets.QMainWindow):#класс главного окна приложения
    def __init__(self):
        global kol
        super(Mywindow, self).__init__()
        self.setupUi()
        self.pushButton_1.clicked.connect(self.ClickButton)
        self.pushButton_2.clicked.connect(self.ClB_l)
        #нажатия кнопок связываются с некоторыми процедурами класса
        self.checkBox_1.stateChanged.connect(self.clickBox_1)
        self.checkBox_2.stateChanged.connect(self.clickBox_2)
        self.checkBox_3.stateChanged.connect(self.clickBox_3)
        self.checkBox_4.stateChanged.connect(self.clickBox_4)
        self.checkBox_5.stateChanged.connect(self.clickBox_5)
        self.checkBox_6.stateChanged.connect(self.clickBox_6)
        self.checkBox_7.stateChanged.connect(self.clickBox_7)
        self.checkBox_8.stateChanged.connect(self.clickBox_8)
        self.checkBox_9.stateChanged.connect(self.clickBox_9)
        self.checkBox_10.stateChanged.connect(self.clickBox_10)
        self.checkBox_11.stateChanged.connect(self.clickBox_11)
        self.checkBox_12.stateChanged.connect(self.clickBox_12)
        self.checkBox_13.stateChanged.connect(self.clickBox_13)
        self.checkBox_14.stateChanged.connect(self.clickBox_14)
        self.checkBox_15.stateChanged.connect(self.clickBox_15)
        self.checkBox_16.stateChanged.connect(self.clickBox_16)
        self.checkBox_17.stateChanged.connect(self.clickBox_17)
        self.checkBox_18.stateChanged.connect(self.clickBox_18)
        self.checkBox_19.stateChanged.connect(self.clickBox_19)
        self.checkBox_20.stateChanged.connect(self.clickBox_20)
        #изменения состояния checkBox связываются с некоторыми процедурами класса

    def setupUi(self):#настройка атрибутов и параметров окна
        self.resize(1250, 700)
        self.centralwidget =QWidget(self)
        self.centralwidget.setObjectName("centralwidget")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setObjectName(u"tabWidget")
        self.tabWidget.setGeometry(QtCore.QRect(20, 5, 1220, 500))
        self.tab = QWidget()
        self.tab1 = QWidget()
        self.tab.setObjectName(u"tab")
        self.tab1.setObjectName(u"tab1")
        self.label01 = QtWidgets.QLabel(str05,self.tab1)
        self.label01.setObjectName(u"label")
        self.label01.setGeometry(QtCore.QRect(20, 0, 1000, 100))
        self.label01.setFont(QtGui.QFont("Times", 13, QtGui.QFont.Bold))
        self.label02 = QtWidgets.QLabel(str06,self.tab1)
        self.label02.setObjectName(u"label")
        self.label02.setGeometry(QtCore.QRect(20, 15, 1200, 150))
        self.label03 = QtWidgets.QLabel(str07,self.tab1)
        self.label03.setObjectName(u"label")
        self.label03.setGeometry(QtCore.QRect(20, 60, 1200, 145))
        self.label04 = QtWidgets.QLabel(str08,self.tab1)
        self.label04.setObjectName(u"label")
        self.label04.setGeometry(QtCore.QRect(20, 170, 1180, 210))
        self.label05 = QtWidgets.QLabel(str09,self.tab1)
        self.label05.setObjectName(u"label")
        self.label05.setGeometry(QtCore.QRect(20, 380, 1180, 75))
        self.label05.setStyleSheet("QLabel { background-color : yellow;}");
        self.tab2 = QtWidgets.QWidget()
        self.tab2.setObjectName(u"tab2")
        self.label06 = QtWidgets.QLabel(str0,self.tab2)
        self.label06.setObjectName(u"label")
        self.label06.setGeometry(QtCore.QRect(20, 10, 1000, 100))
        self.label06.setFont(QtGui.QFont("Times", 13, QtGui.QFont.Bold))
        self.label07 = QtWidgets.QLabel(str01,self.tab2)
        self.label07.setObjectName(u"label")
        self.label07.setGeometry(QtCore.QRect(20, 50, 1200, 150))
        self.label08 = QtWidgets.QLabel(str03,self.tab2)
        self.label08.setObjectName(u"label")
        self.label08.setGeometry(QtCore.QRect(20, 150, 1200, 150))
        self.label09 = QtWidgets.QLabel(str04,self.tab2)
        self.label09.setObjectName(u"label")
        self.label09.setGeometry(QtCore.QRect(20, 280, 1180, 95))
        self.label09.setStyleSheet("QLabel { background-color : yellow;}");
        self.checkBox_1 = QtWidgets.QCheckBox( "Возраст студента",self.tab)
        self.checkBox_1.setGeometry(QtCore.QRect(150, 50, 300, 18))
        self.checkBox_1.setObjectName("checkBox")
        self.checkBox_2 = QtWidgets.QCheckBox("Наличие/отсуствие высшего образования",self.tab)
        self.checkBox_2.setGeometry(QtCore.QRect(150, 90, 400, 18))
        self.checkBox_2.setObjectName("checkBox_2")
        self.checkBox_3 = QtWidgets.QCheckBox( "Наличие/отсуствие работы",self.tab)
        self.checkBox_3.setGeometry(QtCore.QRect(150, 130, 400, 18))
        self.checkBox_3.setObjectName("checkBox_3")
        self.checkBox_4 = QtWidgets.QCheckBox("Семейное положение",self.tab)
        self.checkBox_4.setGeometry(QtCore.QRect(150, 170, 400, 18))
        self.checkBox_4.setObjectName("checkBox_4")
        self.checkBox_5 = QtWidgets.QCheckBox("Наличие/отсуствие детей",self.tab)
        self.checkBox_5.setGeometry(QtCore.QRect(150, 210, 400, 18))
        self.checkBox_5.setObjectName("checkBox_5")
        self.checkBox_6 = QtWidgets.QCheckBox( "Наличие/отсуствие родителя(ей)",self.tab)
        self.checkBox_6.setGeometry(QtCore.QRect(150, 250, 450, 18))
        self.checkBox_6.setObjectName("checkBox_6")
        self.checkBox_7 = QtWidgets.QCheckBox("Наличие/отсуствие высшего образования у родителя(ей)",self.tab)
        self.checkBox_7.setGeometry(QtCore.QRect(150, 290, 450, 18))
        self.checkBox_7.setObjectName("checkBox_7")
        self.checkBox_8 = QtWidgets.QCheckBox( "Условия проживания",self.tab)
        self.checkBox_8.setGeometry(QtCore.QRect(150, 330, 450, 18))
        self.checkBox_8.setObjectName("checkBox_8")
        self.checkBox_9 = QtWidgets.QCheckBox("Факультет",self.tab)
        self.checkBox_9.setGeometry(QtCore.QRect(150, 370, 450, 18))
        self.checkBox_9.setObjectName("checkBox_9")
        self.checkBox_10 = QtWidgets.QCheckBox("Участие в общественной жизни",self.tab)
        self.checkBox_10.setGeometry(QtCore.QRect(150, 410, 450, 18))
        self.checkBox_10.setObjectName("checkBox_10")
        self.checkBox_11 = QtWidgets.QCheckBox("Среднее количество 3-к за семестр",self.tab)
        self.checkBox_11.setGeometry(QtCore.QRect(700, 50, 450, 18))
        self.checkBox_11.setObjectName("checkBox_11")
        self.checkBox_12 = QtWidgets.QCheckBox("Среднее количество 4-к за семестр",self.tab)
        self.checkBox_12.setGeometry(QtCore.QRect(700, 90, 450, 18))
        self.checkBox_12.setObjectName("checkBox_12")
        self.checkBox_13 = QtWidgets.QCheckBox("Среднее количество 5-к за семестр",self.tab)
        self.checkBox_13.setGeometry(QtCore.QRect(700, 130, 450, 18))
        self.checkBox_13.setObjectName("checkBox_13")
        self.checkBox_14 = QtWidgets.QCheckBox("Среднее количество пересдач за семестр",self.tab)
        self.checkBox_14.setGeometry(QtCore.QRect(700, 170, 460, 18))
        self.checkBox_14.setObjectName("checkBox_14")
        self.checkBox_15 = QtWidgets.QCheckBox( "Количество пересдач с комиссией за все время обучения",self.tab)
        self.checkBox_15.setGeometry(QtCore.QRect(700, 210, 460, 18))
        self.checkBox_15.setObjectName("checkBox_15")
        self.checkBox_16 = QtWidgets.QCheckBox("Количество пропущенных в семестр лекционных занятий",self.tab)
        self.checkBox_16.setGeometry(QtCore.QRect(700, 250, 460, 18))
        self.checkBox_16.setObjectName("checkBox_16")
        self.checkBox_17 = QtWidgets.QCheckBox("Количество пропущенных в семестр практических занятий",self.tab)
        self.checkBox_17.setGeometry(QtCore.QRect(700, 290, 470, 18))
        self.checkBox_17.setObjectName("checkBox_17")
        self.checkBox_18 = QtWidgets.QCheckBox("Сроки сдачи учебных работ",self.tab)
        self.checkBox_18.setGeometry(QtCore.QRect(700, 330, 470, 18))
        self.checkBox_18.setObjectName("checkBox_18")
        self.checkBox_19 = QtWidgets.QCheckBox("Количество пересдач за все время обучения",self.tab)
        self.checkBox_19.setGeometry(QtCore.QRect(700, 370, 460, 18))
        self.checkBox_19.setObjectName("checkBox_19")
        self.checkBox_20 = QtWidgets.QCheckBox("Количество сессий, закрытых на степендию",self.tab)
        self.checkBox_20.setGeometry(QtCore.QRect(700, 410, 460, 18))
        self.checkBox_20.setObjectName("checkBox_20")
        self.tabWidget.addTab(self.tab, "Признаки для группирования")
        self.tabWidget.addTab(self.tab1, "Алгоритм К-средних++")
        self.tabWidget.addTab(self.tab2, "Иерархический алгоритм")
        self.pushButton_1 = QtWidgets.QPushButton( "Провести группирование иерархическим алгоритмом",self.centralwidget)
        self.pushButton_1.setGeometry(QtCore.QRect(150, 550, 410, 100))
        self.pushButton_1.setObjectName("pushButton_1")
        self.pushButton_2 = QtWidgets.QPushButton( "Провести группирование алгоритмом k-средних++",self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(700, 550, 410, 100))
        self.pushButton_2.setObjectName("pushButton_2")
        self.setCentralWidget(self.centralwidget)
        self.setWindowTitle(QtCore.QCoreApplication.translate("Errordiolog", "Группирование студентов", None))      
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(self)    

    def proverka(self):
        #Функция проверки выбора пользователем хотя бы одного признака
         if (self.checkBox_1.isChecked()==False and self.checkBox_2.isChecked()==False
            and self.checkBox_3.isChecked()==False and self.checkBox_4.isChecked()==False
            and self.checkBox_5.isChecked()==False and self.checkBox_6.isChecked()==False
            and self.checkBox_7.isChecked()==False and self.checkBox_8.isChecked()==False
            and self.checkBox_9.isChecked()==False and self.checkBox_10.isChecked()==False
            and self.checkBox_11.isChecked()==False and self.checkBox_12.isChecked()==False
            and self.checkBox_13.isChecked()==False and self.checkBox_14.isChecked()==False
            and self.checkBox_15.isChecked()==False and self.checkBox_16.isChecked()==False
            and self.checkBox_17.isChecked()==False and self.checkBox_18.isChecked()==False
            and self.checkBox_19.isChecked()==False and self.checkBox_20.isChecked()==False):
             #Возвращает True в случае выбора хотя бы одного признака; 
             #в ином случае возвращает False
             return False 
         else:
             return True

    def ClB_l(self):
        #срабатывает при нажатии на кнопку "Провести анализ алгоритмом k-средних++"
        global L
        print(L)
        if self.proverka()==False:
        #если не выбрано признаков-уведомление об ошибке
         err.show()  
         return
        #ошибки нет-переход к вводу m количества кластеров
        test.show()
        #отображение окна для ввода количества кластеров
        #сохранение структуры анализируемых данных
        
    def ClickButton(self):
        #срабатывает при нажатии на кнопку "Провести анализ Иерархическим алгоритмом"
        if self.proverka()==False:
        #если не выбрано признаков-уведомление об ошибке
         err.show()  
         return
        #ошибки нет-переход к иерархическому методу кластерного анализа
        self.Ier_a()
    
    def Ier_a(self):
        #Процедура, реализующая алгоритм иерархического кластерного анализа
        global srn,std,figu,canvas,Z,test2
        _do=data1()
        #загружаются значения признаков без идентификаторов студентов
        srn=_do
         #массив значений запоминается
        _do=(_do-_do.mean(axis=0))/_do.std(axis=0)
        #данные стандартизируются
        std=_do
        #сохранение стандартизированных данных
        from scipy.cluster.hierarchy import linkage,fcluster,dendrogram
        Z=linkage(_do,method='ward',metric='euclidean') 
        #матрица связи
        figu=matplotlib.pyplot.figure()
        dendogr=dendrogram(Z,orientation='top',labels=data.ID_студента.values)
        #дендрограмма
        test2.verticalLayout.removeWidget(canvas)
        canvas=Mycanvas(figu)
        test2.verticalLayout.addWidget(canvas)
        #отображение дендрограммы в окне
        print("шозафигня")
        test2.show()

    def clickBox_1(self):
       global L
       if self.checkBox_1.isChecked():                     
            L.append("Возраст") 
            #если признак выбран, то он добавляется в список L   
       else:
            global data_two
            L.remove("Возраст")
            #если признак не выбран, то он удаляется из L 
           
    def clickBox_2(self):
       global L
       if self.checkBox_2.isChecked():            
            L.append('Первое_высшее_образование') 
            #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Первое_высшее_образование')
            #если признак не выбран, то он удаляется из L 
           
    def clickBox_3(self):
       global L
       if self.checkBox_3.isChecked():           
            L.append('Работа')
            #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Работа')
            #если признак не выбран, то он удаляется из L 
    
    def clickBox_4(self):
       global L
       if self.checkBox_4.isChecked():           
            L.append('Семейное_положение')
            #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Семейное_положение')
            #если признак не выбран, то он удаляется из L 

    def clickBox_5(self):
       global L
       if self.checkBox_5.isChecked():           
            L.append('Дети')
            #если признак выбран, то он добавляется в список L   
       else:      
            L.remove('Дети')
            #если признак не выбран, то он удаляется из L 

    def clickBox_6(self):
       global L
       if self.checkBox_6.isChecked():            
            L.append('Оба_родителя')
            #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Оба_родителя')
            #если признак не выбран, то он удаляется из L 

    def clickBox_7(self):
       global L
       if self.checkBox_7.isChecked():            
            L.append('Наличие_образования_у_родителей')
            #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Наличие_образования_у_родителей')
            #если признак не выбран, то он удаляется из L 

    def clickBox_8(self):
      global L     
      if self.checkBox_8.isChecked():            
            L.append('Условия_проживания')           
            #если признак выбран, то он добавляется в список L   
      else:
            L.remove('Условия_проживания')
            #если признак не выбран, то он удаляется из L 

    def clickBox_9(self):
       global L
       if self.checkBox_9.isChecked():
            L.append('Факультет')          
            #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Факультет')
            #если признак не выбран, то он удаляется из L 
           
    def clickBox_10(self):
       global L
       if self.checkBox_10.isChecked():
            L.append('Участие_в_общественной_жизни_университета') 
            #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Участие_в_общественной_жизни_университета')
            #если признак не выбран, то он удаляется из L 

    def clickBox_11(self):
       global L
       if self.checkBox_11.isChecked():  
           L.append('Среднее_количество_3-к')
           #если признак выбран, то он добавляется в список L   
       else:
           L.remove('Среднее_количество_3-к')
           #если признак не выбран, то он удаляется из L 

    def clickBox_12(self):
       global L
       if self.checkBox_12.isChecked():
            L.append('Среднее_количество_4-к')
            #если признак выбран, то он добавляется в список L   
       else: 
            L.remove('Среднее_количество_4-к')
            #если признак не выбран, то он удаляется из L 

    def clickBox_13(self):
       global L
       if self.checkBox_13.isChecked():
            L.append('Среднее_количество_5-к')        
            #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Среднее_количество_5-к')
            #если признак не выбран, то он удаляется из L 

    def clickBox_14(self):
       global L
       if self.checkBox_14.isChecked():
            L.append('Среднее_количество_передач_за_семестр')
            #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Среднее_количество_передач_за_семестр')
            #если признак не выбран, то он удаляется из L 

    def clickBox_15(self):
       global L
       if self.checkBox_15.isChecked():
           L.append('Количество_пересдач_с_коммиссией')          
           #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Количество_пересдач_с_коммиссией')
            #если признак не выбран, то он удаляется из L 

    def clickBox_16(self):
       global L
       if self.checkBox_16.isChecked():
            L.append('Семестровое_количество_пропущенных_лекций')          
            #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Семестровое_количество_пропущенных_лекций')
            #если признак не выбран, то он удаляется из L 

    def clickBox_17(self):
       global L
       if self.checkBox_17.isChecked():
            L.append('Семестровое_количество_пропущенных_занятий')         
            #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Семестровое_количество_пропущенных_занятий')
            #если признак не выбран, то он удаляется из L 

    def clickBox_18(self):
       global L
       if self.checkBox_18.isChecked():
            L.append('Сроки_сдачи_учебных_работ')
            #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Сроки_сдачи_учебных_работ')
            #если признак не выбран, то он удаляется из L 

    def clickBox_19(self):
       global L
       if self.checkBox_19.isChecked():
            L.append('Среднее_количество_пересдач_за_все_время_обучения')  
            #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Среднее_количество_пересдач_за_все_время_обучения')        
            #если признак не выбран, то он удаляется из L 

    def clickBox_20(self):
       if self.checkBox_20.isChecked():
            L.append('Количество_сессий_закрытых_на_степендию')    
            #если признак выбран, то он добавляется в список L   
       else:
            L.remove('Количество_сессий_закрытых_на_степендию')
            #если признак не выбран, то он удаляется из L 

class Myvivod(QtWidgets.QWidget):#класс окна для вывода результатов
    #str передается в заголовок окна
    def __init__(self,str):
        super(Myvivod, self).__init__()
        self.setupUi(str)
        if (str=="Алгоритм к-средних++"):
         self.b1.clicked.connect(self.ClickButton)
        if (str=="Иерархический алгоритм"):
         self.b1.clicked.connect(self.ClickButton1)
        #при использовании разных методов используются разные
        #алгоритмы классификации 
        self.b2.clicked.connect(self.sohr)
        
    def setupUi(self,str):#настройка атрибутов и параметров окна
        self.resize(1250, 900)
        self.tab_1=QtWidgets.QTabWidget(self)
        self.tab_1.setWindowTitle(str)
        self.tab_1.setGeometry(QtCore.QRect(20, 5, 1220, 500))
        self.vvod_d=QWidget(self)
        self.vvod_d.setGeometry(QtCore.QRect(20, 550, 1220, 150))
        self.layout_d = QtWidgets.QVBoxLayout(self.vvod_d)
        self.table1 = QtWidgets.QTableWidget(1,21)
        self.table1.setHorizontalHeaderLabels(data.columns)
        self.table1.resizeColumnsToContents()
        from openpyxl import load_workbook
        book = load_workbook("Answers.xlsx")
        b1 = book.get_sheet_by_name("answers")
        i = 1
        while b1.cell(row=i, column=1).value is not None:
         i=i+1
        #автоматическая запись идентификатора студента в таблицу введения данных
        #о новом студенте
        self.table1.setItem(0,0, QtWidgets.QTableWidgetItem("Студент №{}".format(i-1)))
        #таблица для введения информации о новом студенте
        self.layout_d.addWidget(self.table1)
        self.b1=QtWidgets.QPushButton('Прогноз',self)
        self.b1.setGeometry(QtCore.QRect(230, 750, 150, 60)) 
        self.b2=QtWidgets.QPushButton('Сохранение',self)
        self.b2.setGeometry(QtCore.QRect(30, 750, 150, 60))
        self.setWindowTitle(QtCore.QCoreApplication.translate("Errordiolog", str, None))  
        self.label = QtWidgets.QLabel('Поле для записи информации о новом студенте',self)
        self.label.setGeometry(QtCore.QRect(30, 510, 415, 60))
        self.label1 = QtWidgets.QLabel('Группа',self)   
        self.label1.setGeometry(QtCore.QRect(450, 750, 95, 60))
        self.label1.setStyleSheet("QLabel { background-color : yellow;}");
        self.cb1 = QtWidgets.QComboBox()
        self.cb1.addItems(["больше 30","15-19", "20-25"])
        self.table1.setCellWidget(0,1,self.cb1)
        self.cb2 = QtWidgets.QComboBox()
        self.cb2.addItems(["да", "нет"])
        self.table1.setCellWidget(0,2,self.cb2)
        self.cb3 = QtWidgets.QComboBox()
        self.cb3.addItems(["да", "нет"])
        self.table1.setCellWidget(0,3,self.cb3)
        self.cb4 = QtWidgets.QComboBox()
        self.cb4.addItems(["женат/замужем", "не женат/не замужем"])
        self.table1.setCellWidget(0,4,self.cb4)
        self.cb5 = QtWidgets.QComboBox()
        self.cb5.addItems(["да", "нет"])
        self.table1.setCellWidget(0,5,self.cb5)
        self.cb6 = QtWidgets.QComboBox()
        self.cb6.addItems(["нет, я рос с одним родителем","да","в основном с одним, т.к. второго видел нечасто","нет, я рос без родителей"])
        self.table1.setCellWidget(0,6,self.cb6)
        self.cb7 = QtWidgets.QComboBox()
        self.cb7.addItems(["нет высшего образования, но есть среднее профессиональное","есть высшее образование", "нет высшего образования"])
        self.table1.setCellWidget(0,7,self.cb7)
        self.cb8 = QtWidgets.QComboBox()
        self.cb8.addItems(["вместе со своими родителями","в общежитии", "в отдельной квартире"])
        self.table1.setCellWidget(0,8,self.cb8)
        self.cb9 = QtWidgets.QComboBox()
        self.cb9.addItems(["АВТИ","ЭнМИ", "ИПЭЭФ", "ИЭТ", "ИЭЭ", "ИРЭ", "ГПИ", "ИнЭИ"\
            , "ИГВИЭ", "ИТАЭ"])
        self.table1.setCellWidget(0,9,self.cb9)
        self.cb10 = QtWidgets.QComboBox()
        self.cb10.addItems(["довольно часто","редко", "время от времени"])
        self.table1.setCellWidget(0,10,self.cb10)
        self.cb11 = QtWidgets.QSpinBox()
        self.cb11.setMinimum(0)
        self.cb11.setMaximum(10)
        self.cb11.setSingleStep(1)
        self.table1.setCellWidget(0,11,self.cb11)
        self.cb12 = QtWidgets.QSpinBox()
        self.cb12.setMinimum(0)
        self.cb12.setMaximum(10)
        self.cb12.setSingleStep(1)
        self.table1.setCellWidget(0,12,self.cb12)
        self.cb13 = QtWidgets.QSpinBox()
        self.cb13.setMinimum(0)
        self.cb13.setMaximum(10)
        self.cb13.setSingleStep(1)
        self.table1.setCellWidget(0,13,self.cb13)
        self.cb14 = QtWidgets.QSpinBox()
        self.cb14.setMinimum(0)
        self.cb14.setMaximum(5)
        self.cb14.setSingleStep(1)
        self.table1.setCellWidget(0,14,self.cb14)
        self.cb15 = QtWidgets.QSpinBox()
        self.cb15.setMinimum(0)
        self.cb15.setMaximum(5)
        self.cb15.setSingleStep(1)
        self.table1.setCellWidget(0,15,self.cb15)
        self.cb16 = QtWidgets.QComboBox()
        self.cb16.addItems(["0-3","4-7", "8-15","более 15"])
        self.table1.setCellWidget(0,16,self.cb16)
        self.cb17 = QtWidgets.QComboBox()
        self.cb17.addItems(["0-3","4-7", "8-15","более 15"])
        self.table1.setCellWidget(0,17,self.cb17)
        self.cb18 = QtWidgets.QComboBox()
        self.cb18.addItems(["скорее сдаю работы в течении семестра , как и предусмотрено учебным планом",\
            "скорее часть работ сдаю в течении семестра, часть оставляю на конец",\
           "скорее сдаю все заранее","все работы сдаю в конце семестра"])
        self.table1.setCellWidget(0,18,self.cb18)
        self.cb19 = QtWidgets.QSpinBox()
        self.cb19.setMinimum(0)
        self.cb19.setMaximum(10)
        self.cb19.setSingleStep(1)
        self.table1.setCellWidget(0,19,self.cb19)
        self.cb20 = QtWidgets.QSpinBox()
        self.cb20.setMinimum(0)
        self.cb20.setMaximum(6)
        self.cb20.setSingleStep(1)
        self.table1.setCellWidget(0,20,self.cb20)

    def ClickButton(self):
           #классификация для метода к-средних 
           _do=self.chit()
           global srn,std,centers
           _do=(_do-srn.mean(axis=0))/srn.std(axis=0)
           self.funk_k(_do,centers)

    def ClickButton1(self):
           #классификация для иерархического метода 
           _do=self.chit()
           global srn,std,labs
           _do=(_do-srn.mean(axis=0))/srn.std(axis=0)
           self.funk_i(_do,std,labs)
    
    def chit(self):
        #считывание данных о новом студенте
        #преобразование данных к количественному виду
        global L 
        text0 =self.table1.item(0,0).text()
        dat_stud = pandas.DataFrame({'ID_студента': [text0]})
        text1 = self.table1.cellWidget(0,1).currentText()
        text2 = self.table1.cellWidget(0,2).currentText()
        text3 = self.table1.cellWidget(0,3).currentText()
        text4 = self.table1.cellWidget(0,4).currentText()
        text5 = self.table1.cellWidget(0,5).currentText()
        text6 = self.table1.cellWidget(0,6).currentText()
        text7 = self.table1.cellWidget(0,7).currentText()
        text8 = self.table1.cellWidget(0,8).currentText()
        text9 = self.table1.cellWidget(0,9).currentText()
        text10 = self.table1.cellWidget(0,10).currentText()
        text11 = self.table1.cellWidget(0,11).value()
        text12 = self.table1.cellWidget(0,12).value()
        text13 = self.table1.cellWidget(0,13).value()
        text14 = self.table1.cellWidget(0,14).value()
        text15 = self.table1.cellWidget(0,15).value()
        text16 = self.table1.cellWidget(0,16).currentText()
        text17 = self.table1.cellWidget(0,17).currentText()
        text18 = self.table1.cellWidget(0,18).currentText()
        text19 = self.table1.cellWidget(0,19).value()
        text20 = self.table1.cellWidget(0,20).value()

        if 'Возраст'  in L:
           dat_stud['15-19']=[0]
           dat_stud['20-25']=[0]
           dat_stud['больше 30']=[0]
        if 'Первое_высшее_образование'  in L:
           dat_stud['Первое высшее образование(0-да,1-нет)']=[0]
        if 'Работа'  in L:
             dat_stud['Работа(1-есть,0-нет)']=[0]
        if 'Семейное_положение'  in L:
            dat_stud['1-не женат/не замужем 0-женат/замужем']=[0]
        if 'Дети'  in L:
            dat_stud['Есть дети (0-да 1-нет)']=[0]          
        if 'Оба_родителя'  in L:
            dat_stud['в основном с одним, т.к. второго видел нечасто']=[0]
            dat_stud['да']=[0] 
            dat_stud['нет, я рос без родителей']=[0]  
            dat_stud['нет, я рос с одним родителем']=[0]  
        if 'Наличие_образования_у_родителей'  in L:
            dat_stud['есть высшее образование']=[0]
            dat_stud['нет высшего образования']=[0]
            dat_stud['нет высшего образования, но есть среднее профессиональное']=[0]           
        if 'Условия_проживания'  in L:
            dat_stud['в общежитии']=[0]
            dat_stud['в отдельной квартире']=[0]
            dat_stud['вместе со своими родителями']=[0]
        if 'Факультет'  in L:
            dat_stud['АВТИ']=[0]
            dat_stud['ГПИ']=[0]
            dat_stud['ИГВИЭ']=[0]
            dat_stud['ИПЭЭФ']=[0]
            dat_stud['ИРЭ']=[0]
            dat_stud['ИТАЭ']=[0]
            dat_stud['ИЭТ']=[0]
            dat_stud['ИЭЭ']=[0]
            dat_stud['ИнЭИ']=[0]
            dat_stud['ЭнМИ']=[0]   
        if 'Участие_в_общественной_жизни_университета'  in L:
            dat_stud['время от времени']=[0]
            dat_stud['довольно часто']=[0]
            dat_stud['редко']=[0]            
        if 'Среднее_количество_3-к'  in L:
            dat_stud['Среднее_количество_3-к']=[0]
        if 'Среднее_количество_4-к'  in L:
            dat_stud['Среднее_количество_4-к']=[0]
        if 'Среднее_количество_5-к'  in L:
            dat_stud['Среднее_количество_5-к']=[0]
        if 'Среднее_количество_передач_за_семестр'  in L:
           dat_stud['Среднее_количество_передач_за_семестр']=[0]
        if 'Количество_пересдач_с_коммиссией'  in L:
            dat_stud['Количество_пересдач_с_коммиссией']=[0]
        if 'Семестровое_количество_пропущенных_лекций'  in L:
            dat_stud['0-3']=[0]
            dat_stud['4-7']=[0]
            dat_stud['8-15']=[0]
            dat_stud['более 15']=[0]
        if  'Семестровое_количество_пропущенных_занятий' in L:
            dat_stud['a']=[0]
            dat_stud['b']=[0]
            dat_stud['c']=[0]
            dat_stud['d']=[0]
        if  'Сроки_сдачи_учебных_работ' in L:
            dat_stud['все работы сдаю в конце семестра']=[0]
            dat_stud['скорее сдаю все заранее']=[0]
            dat_stud['скорее сдаю работы в течении семестра , как и предусмотрено учебным планом']=[0]
            dat_stud['скорее часть работ сдаю в течении семестра, часть оставляю на конец']=[0]
        if  'Среднее_количество_пересдач_за_все_время_обучения' in L:
            dat_stud['Среднее_количество_пересдач_за_все_время_обучения']=[0]
        if  'Количество_сессий_закрытых_на_степендию' in L:
            dat_stud['Количество_сессий_закрытых_на_степендию']=[0]
               
        if text1  in dat_stud.columns:
            dat_stud[text1]=1
        if 'Первое высшее образование(0-да,1-нет)'  in dat_stud.columns:
            if  (text2=='нет'):
             dat_stud['Первое высшее образование(0-да,1-нет)']=1
        if 'Работа(1-есть,0-нет)'  in dat_stud.columns:
            if  (text3=='да'):
             dat_stud['Работа(1-есть,0-нет)']=1
        if  '1-не женат/не замужем 0-женат/замужем' in dat_stud.columns: 
            if  (text4=='не женат/не замужем'):
                dat_stud['1-не женат/не замужем 0-женат/замужем']=1
        if  'Есть дети (0-да 1-нет)' in dat_stud.columns:
            if  (text5=='нет'):
                dat_stud['Есть дети (0-да 1-нет)']=1
        if text6  in dat_stud.columns:
            dat_stud[text6]=1
        if text7  in dat_stud.columns:
            dat_stud[text7]=1
        if text8  in dat_stud.columns:
            dat_stud[text8]=1
        if text9  in dat_stud.columns:
            dat_stud[text9]=1
        if text10  in dat_stud.columns:
            dat_stud[text10]=1
        if  'Среднее_количество_3-к' in dat_stud.columns:
            dat_stud['Среднее_количество_3-к']=text11
        if  'Среднее_количество_4-к' in dat_stud.columns:
            dat_stud['Среднее_количество_4-к']=text12
        if  'Среднее_количество_5-к' in dat_stud.columns:
            dat_stud['Среднее_количество_5-к']=text13
        if  'Среднее_количество_передач_за_семестр' in dat_stud.columns:
            dat_stud['Среднее_количество_передач_за_семестр']=text14
        if  'Количество_пересдач_с_коммиссией' in dat_stud.columns:
            dat_stud['Количество_пересдач_с_коммиссией']=text15
        if text16  in dat_stud.columns:
            dat_stud[text16]=1 
        if  'a' and 'b' and 'c' and 'd' in dat_stud.columns:
           if  (text17=='0-3'):
               dat_stud['a']=1 
           if  (text17=='4-7'):
               dat_stud['b']=1
           if  (text17=='8-15'):
               dat_stud['c']=1
           if  (text17=='более 15'):
               dat_stud['d']=1
        if text18  in dat_stud.columns:
            dat_stud[text18]=1 
        if  'Среднее_количество_пересдач_за_все_время_обучения' in dat_stud.columns:
            dat_stud['Среднее_количество_пересдач_за_все_время_обучения']=text19
        if  'Количество_сессий_закрытых_на_степендию' in dat_stud.columns:
            dat_stud['Количество_сессий_закрытых_на_степендию']=text20
        return dat_stud.iloc[:,1:].values   
   
    def funk_k(self,do,centers):
        #процедура классификации для метода к-средних
        from scipy.spatial import distance
        min=distance.euclidean(centers[0], do)
        m=0
        gr=0
        for clast in centers:
         m=m+1
         d = distance.euclidean(clast, do)
         if (d<min):
          min=d
          gr=m
         if (gr==0):
          gr=1
        self.label1.setText("Группа №{}".format(gr) ) 
      
    def funk_i(self,do,std,labs):  
        #процедура классификации для иерархического метода
        from scipy.spatial import distance
        min=distance.euclidean(std[0], do)
        m=0
        gr=0
        for clast in std:
          d = distance.euclidean(clast, do)
          if (d<min):
           min=d
           gr=labs[m]
          m=m+1
        if (gr==0):
            gr=labs[0]
        self.label1.setText("Группа №{}".format(gr) ) 
         
    def sohr(self): 
        #сохранение данных о новом студене в документ Answers.xlsx
        # сохранение данных о новом студенте в data и data_preob
        from openpyxl import load_workbook
        book = load_workbook("Answers.xlsx")
        b1 = book.get_sheet_by_name("answers")
        i = 1
        while b1.cell(row=i, column=1).value is not None:
         i=i+1
        tex0=self.table1.item(0,0).text()
        b1.cell(row=i, column=1).value = tex0
        tex1=self.table1.cellWidget(0,1).currentText()
        b1.cell(row=i, column=2).value = tex1
        tex2=self.table1.cellWidget(0,2).currentText()
        b1.cell(row=i, column=3).value= tex2
        tex3=self.table1.cellWidget(0,3).currentText()
        b1.cell(row=i, column=4).value = tex3
        tex4=self.table1.cellWidget(0,4).currentText()
        b1.cell(row=i, column=5).value= tex4
        tex5=self.table1.cellWidget(0,5).currentText()
        b1.cell(row=i, column=6).value = tex5
        tex6=self.table1.cellWidget(0,6).currentText()
        b1.cell(row=i, column=7).value = tex6
        tex7=self.table1.cellWidget(0,7).currentText()
        b1.cell(row=i, column=8).value= tex7
        tex8=self.table1.cellWidget(0,8).currentText()
        b1.cell(row=i, column=9).value= tex8
        tex9=self.table1.cellWidget(0,9).currentText()
        b1.cell(row=i, column=10).value = tex9
        tex10=self.table1.cellWidget(0,10).currentText()
        b1.cell(row=i, column=11).value = tex10
        tex11=self.table1.cellWidget(0,11).value()
        b1.cell(row=i, column=12).value= tex11
        tex12=self.table1.cellWidget(0,12).value()
        b1.cell(row=i, column=13).value = tex12
        tex13=self.table1.cellWidget(0,13).value()
        b1.cell(row=i, column=14).value = tex13
        tex14=self.table1.cellWidget(0,14).value()
        b1.cell(row=i, column=15).value = tex14
        tex15=self.table1.cellWidget(0,15).value()
        b1.cell(row=i, column=16).value = tex15
        tex16=self.table1.cellWidget(0,16).currentText()
        b1.cell(row=i, column=17).value = tex16
        tex17=self.table1.cellWidget(0,17).currentText()
        b1.cell(row=i, column=18).value = tex17
        tex18=self.table1.cellWidget(0,18).currentText()
        b1.cell(row=i, column=19).value = tex18
        tex19=self.table1.cellWidget(0,19).value()
        b1.cell(row=i, column=20).value = tex19
        tex20=self.table1.cellWidget(0,20).value()
        b1.cell(row=i, column=21).value = tex20
        book.save("Answers.xlsx")
        global data
        data.loc[i-2]=[tex0,tex1,tex2,tex3,tex4,tex5,tex6,tex7,tex8,tex9,tex10,tex11,tex12,tex13,tex14,tex15,tex16,tex17,tex18,tex19,tex20]
        print(data)
        self.dob(tex0,tex1,tex2,tex3,tex4,tex5,tex6,tex7,tex8,tex9,tex10,tex16,tex17,tex18,i)
        self.table1.setItem(0,0, QtWidgets.QTableWidgetItem("Студент №{}".format(i)))
        sohr.show()

    def dob(self,tex0,tex1,tex2,tex3,tex4,tex5,tex6,tex7,tex8,tex9,tex10,tex16,tex17,tex18,i):
        #функция добавления новых данных в data_preob
        global dat_preob
        dat_preob.loc[i-2]=[tex0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        if tex1=='больше 30':        
         dat_preob['Возраст_больше_30'][i-2]=1 
        if tex1=='15-19':        
         dat_preob['Возраст_15-19'][i-2]=1 
        if tex1=='20-25':        
         dat_preob['Возраст_20-15'][i-2]=1 
        if tex2=='да':
         dat_preob['Первое_высшее_образование_преоб'][i-2]=0
        else:
         dat_preob['Первое_высшее_образование_преоб'][i-2]=1
        if tex3=='да':
         dat_preob['Работа_преоб'][i-2]=1
        else:
         dat_preob['Работа_преоб'][i-2]=0
        if tex4=='не женат/не замужем':
         dat_preob['Семейное_положение_преоб'][i-2]=1
        else:
         dat_preob['Семейное_положение_преоб'][i-2]=0
        if tex5=='да':
         dat_preob['Дети_преоб'][i-2]=0
        else:
         dat_preob['Дети_преоб'][i-2]=1
        if tex6=='нет, я рос с одним родителем':
         dat_preob['Рос_с_одним_родителем'][i-2]=1 
        if tex6=='да':
         dat_preob['Рос_с_обоими_родителями'][i-2]=1 
        if tex6=='в основном с одним, т.к. второго видел нечасто':
         dat_preob['Рос_в_основном_с_одним_родителем'][i-2]=1 
        if tex6=='нет, я рос без родителей':
         dat_preob['Рос_без_родителей'][i-2]=1 
        if tex7=='нет высшего образования, но есть среднее профессиональное':
         dat_preob['У_родителей_есть_среднее_проф_образование'][i-2]=1
        if tex7=='есть высшее образование':
         dat_preob[ 'У_родителей_есть_высшее_образование'][i-2]=1
        if tex7=='нет высшего образования':
         dat_preob['У_родителей_нет_высшего_образования'][i-2]=1
        if tex8=='вместе со своими родителями':  
         dat_preob['Проживание_с_родителями'][i-2]=1
        if tex8=='в общежитии':  
         dat_preob['Проживание_в_общежитии'][i-2]=1
        if tex8=='в отдельной квартире':  
         dat_preob['Проживании_в_отдельной_квартире'][i-2]=1
        if tex9=='АВТИ':
         dat_preob['Факультет_АВТИ'][i-2]=1
        if tex9=='ЭнМИ':
         dat_preob['Факультет_ЭнМи'][i-2]=1
        if tex9=='ИПЭЭФ':
         dat_preob['Факультет_ИПЭЭФ'][i-2]=1
        if tex9=='ИЭТ':
         dat_preob['Факультет_ИЭТ'][i-2]=1
        if tex9=='ИЭЭ':
         dat_preob['Факультет_ИЭЭ'][i-2]=1
        if tex9=='ИРЭ':
         dat_preob['Факультет_ИРЭ'][i-2]=1
        if tex9=='ГПИ':
         dat_preob['Факультет_ГПИ'][i-2]=1
        if tex9=='ИнЭИ':
         dat_preob['Факультет_ИнЭи'][i-2]=1
        if tex9=='ИГВИЭ':
         dat_preob['Факультет_ИГВИЭ'][i-2]=1
        if tex9=='ИТАЭ':
         dat_preob['Факультет_ИТАЭ'][i-2]=1
        if tex10=='довольно часто':
         dat_preob['Участие_в_общественной_ жизни_довольно_частое'][i-2]=1 
        if tex10=='редко':
         dat_preob['Участие_в_общественной_ жизни_редкое'][i-2]=1 
        if tex10=='время от времени':
         dat_preob['Участие_в_общественной_ жизни_время_от_времени'][i-2]=1 
        if tex16=='0-3':
         dat_preob['0-3_пропущенных_лекции'][i-2]=1
        if tex16=='4-7':
         dat_preob['4-7_пропущенных_лекции'][i-2]=1
        if tex16=='8-15':
         dat_preob['8-15_пропущенных_лекции'][i-2]=1
        if tex16=='более 15':
         dat_preob['более_15_пропущенных_лекций'][i-2]=1
        if tex17=='0-3':
         dat_preob['0-3_пропущенных_занятий'][i-2]=1
        if tex17=='4-7':
         dat_preob['4-7_пропущенных_занятий'][i-2]=1
        if tex17=='8-15':
         dat_preob['8-15_пропущенных_занятий'][i-2]=1
        if tex17=='более 15':
         dat_preob['более_15_пропущенных_ занятий'][i-2]=1
        if tex18=='скорее сдаю работы в течении семестра , как и предусмотрено учебным планом':
         dat_preob['Сдача_работ_в_течении_семестра'][i-2]=1 
        if tex18=='скорее часть работ сдаю в течении семестра, часть оставляю на конец':
         dat_preob['Сдача_работ_в_течение_и_ в_конце_семестра'][i-2]=1 
        if tex18=='скорее сдаю все заранее':
         dat_preob['Сдача_работ_заранее'][i-2]=1 
        if tex18=='все работы сдаю в конце семестра':
         dat_preob['Сдача_работ_в_конце_семестра'][i-2]=1 

       
        

if __name__ == "__main__": #начало основной программы
  
 str1="Ошибка №1"
 str2="Вы не выбрали ни одного признака для группирования.\n        Пожалуйста, выберите хотя бы один признак!"
 str3="Алгоритм к-средних++"
 str4="Введите количество m кластеров."
 str5="Ошибка №2"
 str6="  Один или более кластеров могут оказаться пустыми.\n  Пожалуйста, выберите меньшее число m кластеров!"   
 str7="Иерархический алгоритм"
 str8="Выберете порог отсечения дендрограммы для получения разбиения!"
 str9="Иерархический алгоритм"
 str10="Алгоритм к-средних++"
 str11="Сохранение"
 str12="   Вы успешно сохранили данные о новом студенте!"
 #переменные для хранения заголовков или надпесей окон приложения

 Z=0 
 #переменная для хранения матрицы связи
 
 figu=matplotlib.pyplot.figure()
 #переменная для хранения дендрограммы
 
 data = pandas.read_excel('C:\\Users\\Nast\\source\\repos\\Test11\\Answers.xlsx')
 #таблица исходных данных
 
 dannie=LabelBinarizer()
 #класс для кодирования данных

 dat_preob=pandas.DataFrame(data['ID_студента'])
 dat_preob= dat_preob.rename(columns={'ID_студента': 'ID_студента_2'})
 d1=pandas.DataFrame(dannie.fit_transform(data['Возраст']),columns=['Возраст_15-19','Возраст_20-15','Возраст_больше_30'])
 dat_preob=dat_preob.join(d1)
 d1=pandas.DataFrame(dannie.fit_transform(data['Первое_высшее_образование']),columns=['Первое_высшее_образование_преоб'])
 dat_preob=dat_preob.join(d1)
 d1=pandas.DataFrame(dannie.fit_transform(data['Работа']),columns=['Работа_преоб'])
 dat_preob=dat_preob.join(d1)
 d1=pandas.DataFrame(dannie.fit_transform(data['Семейное_положение']),columns=['Семейное_положение_преоб'])
 dat_preob=dat_preob.join(d1)
 d1=pandas.DataFrame(dannie.fit_transform(data['Дети']),columns=['Дети_преоб'])
 dat_preob=dat_preob.join(d1)
 d1=pandas.DataFrame(dannie.fit_transform(data['Оба_родителя']),columns=['Рос_в_основном_с_одним_родителем','Рос_с_обоими_родителями','Рос_без_родителей','Рос_с_одним_родителем'])
 dat_preob=dat_preob.join(d1)
 d1=pandas.DataFrame(dannie.fit_transform(data['Наличие_образования_у_родителей']),columns=['У_родителей_есть_высшее_образование','У_родителей_нет_высшего_образования','У_родителей_есть_среднее_проф_образование'])
 dat_preob=dat_preob.join(d1)
 d1=pandas.DataFrame(dannie.fit_transform(data['Условия_проживания']),columns=['Проживание_в_общежитии','Проживании_в_отдельной_квартире','Проживание_с_родителями'])
 dat_preob=dat_preob.join(d1)
 d1=pandas.DataFrame(dannie.fit_transform(data['Факультет']),columns=['Факультет_АВТИ','Факультет_ГПИ','Факультет_ИГВИЭ','Факультет_ИПЭЭФ','Факультет_ИРЭ','Факультет_ИТАЭ','Факультет_ИЭТ','Факультет_ИЭЭ','Факультет_ИнЭи','Факультет_ЭнМи'])
 dat_preob=dat_preob.join(d1)
 d1=pandas.DataFrame(dannie.fit_transform(data['Участие_в_общественной_жизни_университета']),columns=['Участие_в_общественной_ жизни_время_от_времени','Участие_в_общественной_ жизни_довольно_частое','Участие_в_общественной_ жизни_редкое'])
 dat_preob=dat_preob.join(d1)
 d1=pandas.DataFrame(dannie.fit_transform(data['Семестровое_количество_пропущенных_лекций']),columns=['0-3_пропущенных_лекции','4-7_пропущенных_лекции','8-15_пропущенных_лекции','более_15_пропущенных_лекций'])
 dat_preob=dat_preob.join(d1)
 d1=pandas.DataFrame(dannie.fit_transform(data['Семестровое_количество_пропущенных_занятий']),columns=['0-3_пропущенных_занятий','4-7_пропущенных_занятий','8-15_пропущенных_занятий','более_15_пропущенных_ занятий'])
 dat_preob=dat_preob.join(d1)
 d1=pandas.DataFrame(dannie.fit_transform(data['Сроки_сдачи_учебных_работ']),columns=['Сдача_работ_в_конце_семестра','Сдача_работ_заранее','Сдача_работ_в_течении_семестра','Сдача_работ_в_течение_и_ в_конце_семестра'])
 dat_preob=dat_preob.join(d1)
 #таблица преобразованных данных
 
 L=[]
 #cписок для хранения выбранных признаков

 srn=0
 #переменная для хранения массива значений признаков

 std=0
 #переменная для хранения массива стандартизированных значений признаков

 centers=0
 #переменная для хранения центров масс кластеров

 labs=0
 #переменная для хранения массива меток кластеров

 app = QtWidgets.QApplication([]) 

 err=Error(str1,str2)
 err2=Error(str5,str6)
 sohr=Error(str11,str12)
 #экземпляры класса окна-оповещения  
 
 test=Kol_N_k_sred(str3,str4)
 #экземпляр класса окна для введения m числа кластеров
 
 test2=Dendr(str7,str8)
 #экземпляр класса окна, содержащего дендрограмму

 canvas=Mycanvas(figu)
 #экземпляр класса "холста" для дендрограммы
 
 vivod_d=Myvivod(str9)
 vivod_d1 =Myvivod(str10)
 #экземпляры класса окна для отображения кластеров 
 
 application = Mywindow()
 #экземпляр класса главного окна приложеня
 
 application.show()
 #показать главное окно
 
 sys.exit(app.exec()) #конец основной программы














   # d=data1()
        #d1=d.iloc[:,1:].values
        #d1=(d1-d1.mean(axis=0))/d1.std(axis=0)
        
        #for i in range(len(d)):
        # d.loc[i,1:]=d1[i]
        
        #min=-1
        #gr=0
        #d.loc[:,'label']=labs
        #for group in d.groupby('label'):
          #n=group[1].iloc[:,1:]
          #n=n.drop('label',1) 
          #n=n.values
          #V1=sum(sum((n-n.mean(axis=0))**2))
          #i=group[1].shape[0] 
          #n=group[1].iloc[:,1:]
          #n=n.drop('label',1) 
          #dat=self.chit()
          #dat=(dat-d1.mean(axis=0))/d1.std(axis=0)
          #print(dat)
          #n.loc[len(d)]=dat.ravel()
          #n=n.values 
          #V2=sum(sum((n-n.mean(axis=0))**2))
          #ab=abs(V1-V2)
         # print(ab)
         # if min<0 or ab<=min:
          #    min=ab
         #     gr=group[1]['label'].values[0]
        #
       # self.label1.setText("Группа №{}".format(gr) ) 
       #использовать квадрат евклидова расстояния или нет

